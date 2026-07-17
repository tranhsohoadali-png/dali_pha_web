# -*- coding: utf-8 -*-
"""XUẤT FILE ĐĂNG HÀNG LOẠT SHOPEE (Mass Upload) — thay bước tạo listing tay.

Điền sẵn file mẫu CỦA CHÍNH SHOP (pha/assets/shopee_mass_upload.xlsx) từ các ảnh
sản phẩm đã ghép ở Xưởng ảnh: tên + mô tả + 2 khổ bán (giá/kho/cân nặng/đóng gói)
+ ảnh bìa (URL công khai) -> chủ vào Seller Centre bấm Đăng hàng loạt là xong cả lô.
KHÔNG cần Open API, không cần mật khẩu, không đụng ToS.

Số đo lấy từ listing THẬT của shop (ảnh chụp Seller Centre 2026-07-17) — xem
_exp/shopee_recon.md. Template đã VÁ XML (Shopee ghi activePane sai chuẩn nên
openpyxl gốc không đọc được).
"""
import os
import re
from datetime import datetime

from django.conf import settings
from django.http import JsonResponse

from pha.views import staff_required

_TEMPLATE = os.path.join(os.path.dirname(__file__), 'assets', 'shopee_mass_upload.xlsx')
_SHEET = 'Bản đăng tải'
_ROW0 = 7                    # dữ liệu bắt đầu từ hàng 7 (1-6 là header/hướng dẫn)
_OUT_SUB = 'shopee'

# Giá trị cố định — lấy từ listing thật của shop
_CATEGORY = '101156-Nhà cửa & Đời sống/Trang trí nhà cửa/Khung ảnh & vật trang trí tường'
_VAR_NAME = 'kích thước'     # tên nhóm phân loại 1 (<=14 ký tự)

# ⚠️ THƯƠNG HIỆU: Shopee đòi **MÃ SỐ** thương hiệu, KHÔNG nhận chữ "DALI"
# (upload lỗi "Row 7: Thương hiệu should be valid number"). Sheet tra ID trong file
# mẫu (HiddenShopBrand) RỖNG nên không có ID. Shopee cho phép ĐỂ TRỐNG rồi đặt
# thương hiệu sau bằng công cụ thuộc tính -> để trống cho upload chạy.
# Có ID rồi thì điền vào đây (chỉ SỐ), vd _BRAND_ID = '1234567'.
_BRAND_ID = ''

# CHỈ bán 2 khổ trên Shopee (user chốt). Giá Shopee = giá web + 100k (rồi Chương
# trình Shop giảm về đúng giá web: 299->199, 399->299 — mass upload chỉ đặt giá gốc).
_SIZES = [
    {'ten': '40x50 cm', 'gia': 299000, 'kho': 95, 'gr': 800,  'dai': 40, 'rong': 50, 'cao': 1},
    {'ten': '50x65 cm', 'gia': 399000, 'kho': 98, 'gr': 1500, 'dai': 65, 'rong': 50, 'cao': 1},
]

# Cột (1-based) trên sheet "Bản đăng tải"
C_CAT, C_NAME, C_DESC, C_SKUP, C_MA = 1, 2, 3, 4, 5
C_VAR1, C_OPT1 = 6, 7
C_PRICE, C_STOCK, C_SKU = 11, 12, 13
C_COVER = 17
C_WEIGHT, C_LEN, C_WID, C_HEI = 26, 27, 28, 29
C_HOATOC, C_NHANH, C_CONGKENH, C_TRONGNGAY = 30, 31, 32, 33
C_SMARTBOX, C_SPX, C_DIEMNHAN = 34, 35, 36
C_BRAND = 38
C_XUATXU, C_CHATLIEU, C_PHONGCACH = 43, 51, 54      # thuộc tính 100037/100134/100169

_NAME_PREFIX = 'Tranh treo tường tự tô màu số hóa DALI'


def _mo_ta(ten, so_mau):
    """Mô tả theo ĐÚNG khuôn shop đang dùng (Shopee bắt 100–3000 ký tự)."""
    kho_list = '\n'.join('                     ' + s['ten'].replace(' cm', 'cm') for s in _SIZES)
    return (
        '🎨 Tranh tô màu theo số DALI — tự tay tô nên bức tranh của riêng bạn, thư giãn sau '
        'ngày dài và có ngay một tác phẩm trang trí cho không gian sống. Toan đã in sẵn các ô '
        'số, bạn chỉ cần tô đúng màu theo số là hoàn thiện — không cần biết vẽ.\n'
        '🖼️ Tranh treo tường giúp không gian thêm ấn tượng, tăng nét sang trọng, hoàn mỹ và '
        'thể hiện gu thẩm mỹ của gia chủ. Cũng là món quà ý nghĩa dành tặng người thân, bạn bè.\n'
        f'Tranh số hóa DALI xin giới thiệu tới mọi người sản phẩm tranh số hóa {ten}\n'
        f'💎 Kích thước tranh: {_SIZES[0]["ten"].replace(" cm", "cm")}\n{kho_list}\n'
        f'💎 Bộ màu đi kèm : Màu Acrylic {so_mau} màu\n'
        '💎 Toan vẽ: Vải Canvas.\n'
        '💎 Bộ tranh bao gồm: Toan vẽ hoặc vải vẽ, bộ màu, bộ bút, bản nháp.\n'
        '💎 Lưu ý: Sản phẩm là loại chưa vẽ. Khách hàng cần tự tay tô màu để có được tranh như mẫu\n'
        '🔷 CÁCH VẼ:\n'
        'Chuẩn bị cốc nước để rửa bút, khăn khô hoặc giấy ăn để lau khô bút sau khi rửa.\n'
        '1️⃣ Lấy màu trong lọ màu tô lên ô có số tương ứng, nên tô hết màu này rồi mới tô sang '
        'màu khác. Nên tô theo thứ tự đã đánh số sẵn (Từ 1 đến hết) hoặc theo gam màu từ nhạt đến đậm.\n'
        '2️⃣ Khi chuyển màu hoặc nghỉ không tô nữa, bạn nên rửa sạch bút và lau khô để lần sau dùng.\n'
        '3️⃣ Khi nghỉ vẽ, nhớ đóng chặt nắp hộp màu lại.'
    )[:3000]


def _ten_sp(ma, ten_ai):
    """Tên Shopee 10–120 ký tự, theo mẫu shop: '<prefix> <tên tranh> <mã>'."""
    lo = (ten_ai or '').strip()
    if lo:
        # Tên AI thường mở đầu 'Tranh Tô Màu Số Hóa/Hoá ...' -> BỎ để khỏi lặp với
        # prefix ('...tự tô màu số hóa DALI Tranh Tô Màu Số Hóa ...'). Tiếng Việt có
        # 2 lối viết 'hóa' và 'hoá' -> phải bắt cả hai.
        lo = re.sub(r'^\s*tranh\s+t[ôo]\s+m[àa]u\s+s[ốo]\s+h(?:óa|oá|oa)\s*',
                    '', lo, flags=re.I | re.UNICODE).strip(' -–—:')
    base = f'{_NAME_PREFIX} {lo}'.strip() if lo else _NAME_PREFIX
    name = f'{base} {ma}'.strip()
    if len(name) > 120:                       # cắt phần tên tranh, GIỮ mã ở cuối
        keep = 120 - len(ma) - 1
        name = (base[:keep].rstrip(' -–—,') + ' ' + ma).strip()
    return name[:120]


def _outs_for_shopee(ids=None):
    """Ảnh đã ghép -> mỗi MÃ 1 sản phẩm. ids = các bản ghi user TÍCH CHỌN (lọc TRƯỚC
    rồi mới gộp theo mã) -> tích đúng ảnh nào thì lấy ảnh đó; không tích = lấy tất cả
    mã (ảnh mới nhất). Cùng 1 mã tích nhiều ảnh -> lấy ảnh MỚI NHẤT."""
    from pha.product_studio import _out_list
    rows = _out_list()                        # đã sắp mới nhất trước
    if ids:
        want = set(ids)
        rows = [o for o in rows if o.get('id') in want]
    seen, items = set(), []
    for o in rows:
        ma = (o.get('ma') or '').strip()
        img = o.get('img') or o.get('web') or ''
        if not ma or not img or ma.lower() in seen:
            continue
        seen.add(ma.lower())
        items.append(o)
    return items


def build_shopee_file(request, outs):
    """Điền template -> file xlsx trong MEDIA_ROOT/shopee/. Trả (file_url, n_sp)."""
    import openpyxl
    if not os.path.isfile(_TEMPLATE):
        raise RuntimeError('Thiếu file mẫu Shopee (pha/assets/shopee_mass_upload.xlsx).')
    from pha.web_publish import _reg
    reg = _reg()                              # sổ đã đăng web -> lấy TÊN do AI viết
    wb = openpyxl.load_workbook(_TEMPLATE)
    ws = wb[_SHEET]
    r = _ROW0
    n = 0
    for o in outs:
        ma = (o.get('ma') or '').strip()
        so_mau = str(o.get('colors') or '').strip() or '24'
        cover = request.build_absolute_uri(o.get('img') or o.get('web') or '')
        ten = _ten_sp(ma, (reg.get(ma) or {}).get('name'))
        mo_ta = _mo_ta(ten, so_mau)
        for s in _SIZES:                      # MỖI KHỔ 1 DÒNG, trường SP lặp lại
            ws.cell(r, C_CAT, _CATEGORY)
            ws.cell(r, C_NAME, ten)
            ws.cell(r, C_DESC, mo_ta)
            ws.cell(r, C_SKUP, ma)
            ws.cell(r, C_MA, ma)              # khoá gom phân loại
            ws.cell(r, C_VAR1, _VAR_NAME)
            ws.cell(r, C_OPT1, s['ten'])
            ws.cell(r, C_PRICE, s['gia'])
            ws.cell(r, C_STOCK, s['kho'])
            ws.cell(r, C_SKU, f"{ma}-{s['ten'].split()[0]}")
            ws.cell(r, C_COVER, cover)
            ws.cell(r, C_WEIGHT, s['gr'])
            ws.cell(r, C_LEN, s['dai'])
            ws.cell(r, C_WID, s['rong'])
            ws.cell(r, C_HEI, s['cao'])
            ws.cell(r, C_HOATOC, 'Mở')
            ws.cell(r, C_NHANH, 'Mở')
            ws.cell(r, C_CONGKENH, 'Mở')
            ws.cell(r, C_TRONGNGAY, 'Mở')
            ws.cell(r, C_SMARTBOX, 'Tắt')     # quá chiều rộng
            ws.cell(r, C_SPX, 'Tắt')          # quá chiều dài
            ws.cell(r, C_DIEMNHAN, 'Mở')
            if _BRAND_ID:                 # để trống nếu chưa có MÃ SỐ thương hiệu
                ws.cell(r, C_BRAND, _BRAND_ID)
            # 3 thuộc tính này Shopee nhận CHUỖI ("giá trị đề xuất hoặc chuỗi ký tự")
            ws.cell(r, C_XUATXU, 'Trong nước')
            ws.cell(r, C_CHATLIEU, 'Gỗ')
            ws.cell(r, C_PHONGCACH, 'Retro')
            r += 1
        n += 1
    d = os.path.join(settings.MEDIA_ROOT, _OUT_SUB)
    os.makedirs(d, exist_ok=True)
    fn = f'shopee_{datetime.now():%Y-%m-%d_%H-%M-%S}_{n}sp.xlsx'
    wb.save(os.path.join(d, fn))
    return f'/media/{_OUT_SUB}/{fn}', n


@staff_required
def xuat_shopee(request):
    """Xuất file đăng hàng loạt Shopee cho các mã trong 'Ảnh đã ghép'.
    ?ids=out-a,out-b (chọn) hoặc bỏ trống = TẤT CẢ mã (mỗi mã lấy ảnh mới nhất)."""
    try:
        ids = [i.strip() for i in (request.GET.get('ids') or '').split(',') if i.strip()]
        outs = _outs_for_shopee(ids)
        if not outs:
            return JsonResponse({'ok': False, 'error': 'Không có mã nào để xuất '
                                 '(ảnh phải có MÃ mới xuất được).'})
        file_url, n = build_shopee_file(request, outs)
        return JsonResponse({'ok': True, 'file_url': file_url, 'n': n,
                             'rows': n * len(_SIZES)})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)
