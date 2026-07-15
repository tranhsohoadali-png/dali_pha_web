import csv
import json
import os
import uuid
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta

try:
    from zoneinfo import ZoneInfo
    _VN = ZoneInfo('Asia/Ho_Chi_Minh')
except Exception:
    _VN = None

from functools import wraps

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse, HttpResponse, HttpResponseNotFound, FileResponse, Http404
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt

from django.db.models import F, Sum, Count

from pha import mixing
from pha import recipes
from pha.models import ProductionLog, ImageResult, PaintStock

_img_executor = ThreadPoolExecutor(max_workers=2)


def staff_required(view):
    """Chỉ CHỦ (is_staff) mới vào được trang quản lý; nhân viên bị đẩy về /app."""
    @wraps(view)
    @login_required(login_url='/login')
    def wrapped(request, *args, **kwargs):
        if not request.user.is_staff:
            return redirect('/app')
        return view(request, *args, **kwargs)
    return wrapped


def login_view(request):
    if request.user.is_authenticated:
        return redirect('/' if request.user.is_staff else '/app')
    if request.method == 'POST':
        u = authenticate(request, username=request.POST.get('username', '').strip(),
                         password=request.POST.get('password', ''))
        if u is not None:
            login(request, u)
            nxt = request.GET.get('next')
            return redirect(nxt or ('/' if u.is_staff else '/app'))
        messages.error(request, 'Sai tài khoản hoặc mật khẩu.')
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('/login')


@csrf_exempt
@staff_required
def nhan_vien(request):
    """Quản lý tài khoản nhân viên / quản lý (chỉ chủ)."""
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_user':
            uname = request.POST.get('username', '').strip()
            pwd = request.POST.get('password', '')
            is_manager = request.POST.get('is_manager') == 'on'
            if not uname or not pwd:
                messages.error(request, 'Thiếu tên đăng nhập hoặc mật khẩu.')
            elif User.objects.filter(username__iexact=uname).exists():
                messages.error(request, f'Tài khoản "{uname}" đã tồn tại.')
            else:
                u = User.objects.create_user(username=uname, password=pwd)
                u.is_staff = is_manager
                u.save()
                messages.info(request, f'Đã tạo {"quản lý" if is_manager else "nhân viên"}: {uname}.')
        elif action == 'delete_user':
            uname = request.POST.get('username', '')
            if uname == request.user.username:
                messages.error(request, 'Không thể tự xoá tài khoản đang dùng.')
            else:
                User.objects.filter(username=uname, is_superuser=False).delete()
                messages.info(request, f'Đã xoá tài khoản {uname}.')
        elif action == 'reset_pw':
            uname = request.POST.get('username', '')
            pwd = request.POST.get('password', '')
            u = User.objects.filter(username=uname).first()
            if u and pwd:
                u.set_password(pwd)
                u.save()
                messages.info(request, f'Đã đổi mật khẩu cho {uname}.')
        return redirect('/nhan-vien')
    users = User.objects.order_by('-is_superuser', '-is_staff', 'username')
    return render(request, 'nhan_vien.html', {'users': users})


def _low_stock_names():
    return [p.name for p in PaintStock.objects.all()
            if p.low_threshold > 0 and p.stock <= p.low_threshold]


@csrf_exempt
@staff_required
def dashboard(request):
    """Bảng điều khiển: số liệu, biểu đồ, dự báo mua sơn, năng suất."""
    now = _now()
    today = now.strftime('%Y-%m-%d')
    this_month = now.strftime('%Y-%m')

    today_n = ProductionLog.objects.filter(day=today).count()
    month_n = ProductionLog.objects.filter(month=this_month).count()
    month_cost = ProductionLog.objects.filter(month=this_month).aggregate(s=Sum('cost'))['s'] or 0

    # 30 ngày qua: số mẻ mỗi ngày
    days = [now.date() - timedelta(days=i) for i in range(29, -1, -1)]
    dc = dict(ProductionLog.objects.filter(day__gte=days[0].strftime('%Y-%m-%d'))
              .values_list('day').annotate(n=Count('id')))
    day_labels = [d.strftime('%d/%m') for d in days]
    day_data = [dc.get(d.strftime('%Y-%m-%d'), 0) for d in days]

    # 12 tháng: chi phí
    months, y, m = [], now.year, now.month
    for _ in range(12):
        months.append(f'{y:04d}-{m:02d}')
        m -= 1
        if m == 0:
            m = 12; y -= 1
    months = months[::-1]
    mc = dict(ProductionLog.objects.values_list('month').annotate(c=Sum('cost')))
    month_labels = [_fmt_month(mm) for mm in months]
    month_cost_data = [round(mc.get(mm, 0) or 0) for mm in months]

    # màu dùng nhiều tháng này
    _, top_rows = _stats('month', this_month)

    # dự báo mua sơn (theo 30 ngày qua)
    since = days[0].strftime('%Y-%m-%d')
    usage30 = defaultdict(float)
    for log in ProductionLog.objects.filter(day__gte=since):
        for c in (log.components or []):
            try:
                usage30[c['name']] += float(c['grams'])
            except (KeyError, TypeError, ValueError):
                pass
    forecast = []
    for b in mixing.get_bases():
        ps, _ = PaintStock.objects.get_or_create(name=b['name'])
        used = usage30.get(b['name'], 0)
        per_day = used / 30.0
        days_left = round(ps.stock / per_day) if per_day > 0 else None
        forecast.append({
            'name': b['name'], 'rgb': b['rgb'], 'stock': round(ps.stock, 1), 'used30': round(used),
            'per_day': round(per_day, 1), 'days_left': days_left,
            'suggest': max(0, round(used - ps.stock)),   # đủ dùng ~30 ngày tới
            'low': ps.low_threshold > 0 and ps.stock <= ps.low_threshold,
        })

    # Gắn ô màu + thanh tỉ lệ cho bảng "màu dùng nhiều"
    base_rgb = {b['name']: b['rgb'] for b in mixing.get_bases()}
    max_g = max((r['grams'] for r in top_rows), default=0) or 1
    for r in top_rows:
        r['rgb'] = base_rgb.get(r['name'])
        r['pct'] = round(r['grams'] / max_g * 100)

    users = list(ProductionLog.objects.filter(month=this_month).values('user')
                 .annotate(n=Count('id'), c=Sum('cost')).order_by('-n'))

    return render(request, 'dashboard.html', {
        'today_n': today_n, 'month_n': month_n, 'month_cost': round(month_cost),
        'low_stock': _low_stock_names(),
        'day_labels': json.dumps(day_labels), 'day_data': json.dumps(day_data),
        'month_labels': json.dumps(month_labels), 'month_cost_data': json.dumps(month_cost_data),
        'top_rows': top_rows, 'forecast': forecast, 'users': users,
        'month_label': _fmt_month(this_month), 'today_label': now.strftime('%d/%m/%Y'),
    })


def _api_key():
    """Khoá API kế toán/lương: ưu tiên đặt trong DB (AppSetting, sửa qua web),
    nếu trống thì dùng biến môi trường KETOAN_API_KEY."""
    try:
        from pha.models import AppSetting
        k = (AppSetting.get('KETOAN_API_KEY', '') or '').strip()
        if k:
            return k
    except Exception:
        pass
    return getattr(settings, 'KETOAN_API_KEY', '')


def _ketoan_cfg():
    """Cấu hình KÉO lương từ phần mềm kế toán (ketoan.tranhdali.vn) về app này."""
    from pha.models import AppSetting
    return {
        'url': (AppSetting.get('KETOAN_SALARY_URL', '') or '').strip(),
        'key': (AppSetting.get('KETOAN_PULL_KEY', '') or '').strip(),
    }


def _ketoan_salary(username, day, month, want_raw=False):
    """Gọi API của phần mềm kế toán để lấy lương 1 nhân viên (theo ngày công đã setup bên đó).

    Hợp đồng API (ketoan.tranhdali.vn tự code 1 endpoint GET, ví dụ /api/luong-nhan-vien):
        GET {KETOAN_SALARY_URL}?key=...&user=<username>&day=YYYY-MM-DD&month=YYYY-MM
        -> {"ok": true, "day": <lương ngày>, "month": <tổng lương tháng>, "month_label": "06/2026"}

    Trả dict {'day','month','month_label'} hoặc None nếu chưa cấu hình / lỗi / hết giờ chờ.
    Linh hoạt nhận nhiều tên trường (day/day_total/luong_ngay, month/total/luong_thang)."""
    cfg = _ketoan_cfg()
    if not cfg['url']:
        return None
    import json as _json
    from urllib.parse import urlencode
    from urllib.request import urlopen, Request
    params = {'user': username or '', 'day': day or '', 'month': month or ''}
    if cfg['key']:
        params['key'] = cfg['key']
    sep = '&' if ('?' in cfg['url']) else '?'
    url = cfg['url'] + sep + urlencode(params)
    try:
        req = Request(url, headers={'Accept': 'application/json',
                                    'User-Agent': 'dali-mau/1.0'})
        with urlopen(req, timeout=5) as resp:
            raw = resp.read().decode('utf-8', 'replace')
        data = _json.loads(raw)
    except Exception as e:
        return {'error': str(e)[:140]}
    if not isinstance(data, dict):
        return {'error': 'Kế toán trả về không đúng JSON object.'}
    if data.get('ok') is False:
        return {'error': str(data.get('error') or data.get('msg') or 'Kế toán báo lỗi.')[:140]}

    def pick(*keys):
        for k in keys:
            v = data.get(k)
            if v is None:
                continue
            if isinstance(v, bool):
                return None
            if isinstance(v, (int, float)):
                return float(v)
            # Chuỗi tiền VND: GIỮ CHỈ CHỮ SỐ (bỏ đ/d/₫/VND, dấu chấm-phẩy nghìn, khoảng trắng,
            # mọi ký tự khác) -> chống lỗi khi kế toán trả "634.615đ" / "634615 d" / "634,615₫".
            import re as _re
            s = _re.sub(r'\D', '', str(v))
            if not s:
                return None
            try:
                return float(s)
            except ValueError:
                return None
        return None
    d = pick('day', 'day_total', 'today', 'luong_ngay', 'luong_hom_nay')
    m = pick('month', 'month_total', 'total', 'luong_thang', 'thang', 'tong', 'monthly',
             'total_month', 'month_salary', 'tong_luong', 'sum', 'thu_nhap_thang', 'luong_thang_nay')
    if d is None and m is None:
        err = {'error': 'Không thấy trường lương (day/month) trong phản hồi.'}
        if want_raw:
            err['raw'] = data
        return err
    res = {'day': round(d or 0), 'month': round(m or 0),
           'month_label': str(data.get('month_label') or '')}
    if want_raw:
        res['raw'] = data                    # để endpoint Test soi KẾ TOÁN trả về gì
    return res


@csrf_exempt
@staff_required
def ketoan_luong_test(request):
    """Quản lý bấm Test: thử kéo lương 1 nhân viên từ kế toán để kiểm tra kết nối."""
    user = (request.GET.get('user') or request.user.username or '').strip()
    now = _now()
    res = _ketoan_salary(user, now.strftime('%Y-%m-%d'), now.strftime('%Y-%m'), want_raw=True)
    cfg = _ketoan_cfg()
    if not cfg['url']:
        return JsonResponse({'ok': False, 'msg': 'Chưa nhập URL API kế toán.'})
    if res is None:
        return JsonResponse({'ok': False, 'msg': 'Chưa cấu hình.'})
    if 'error' in res:
        return JsonResponse({'ok': False, 'user': user, 'msg': res['error'],
                             'raw': res.get('raw'), 'fields': list((res.get('raw') or {}).keys())})
    return JsonResponse({'ok': True, 'user': user, 'day': res['day'], 'month': res['month'],
                         'month_label': res.get('month_label', ''),
                         'raw': res.get('raw'), 'fields': list((res.get('raw') or {}).keys())})


@csrf_exempt
def api_xu_ly_anh(request):
    """API xử lý ảnh cho web bán hàng (tranhdali.vn/thiet-ke).
    Nhận 1 ảnh (multipart field 'image') + thông số -> chạy tăng cường AI +
    bản đồ màu ĐỒNG BỘ (đợi xong) -> trả JSON: URL ảnh kết quả + bảng màu.
    Bảo vệ bằng khoá API: header X-API-Key hoặc ?key= (THIETKE_API_KEY).
    """
    from pha.models import AppSetting

    def _cors(resp):
        resp['Access-Control-Allow-Origin'] = getattr(settings, 'THIETKE_ALLOW_ORIGIN', '*')
        resp['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        resp['Access-Control-Allow-Headers'] = 'Content-Type, X-API-Key'
        resp['Cache-Control'] = 'no-store'
        return resp

    if request.method == 'OPTIONS':
        return _cors(HttpResponse(status=204))
    if request.method != 'POST':
        return _cors(JsonResponse({'ok': False, 'error': 'POST only'}, status=405))

    want_key = (AppSetting.get('THIETKE_API_KEY', '') or getattr(settings, 'THIETKE_API_KEY', '')).strip()
    got_key = (request.headers.get('X-API-Key') or request.GET.get('key')
               or request.POST.get('key') or '').strip()
    if not want_key or got_key != want_key:
        return _cors(JsonResponse({'ok': False, 'error': 'Sai hoặc thiếu khoá API'}, status=401))

    upload = request.FILES.get('image')
    if not upload:
        return _cors(JsonResponse({'ok': False, 'error': 'Thiếu ảnh (field image)'}, status=400))

    from pha.imageproc import process_image, split_list

    enhance = (request.POST.get('enhance', '1') in ('1', 'on', 'true'))
    preset_key = (request.POST.get('preset') or 'anime').strip()
    try:
        color_limit = max(0, min(int(request.POST.get('color_limit') or 0), 250))
    except ValueError:
        color_limit = 0
    size_str = (request.POST.get('print_size') or '40x50').strip()
    try:
        dims = [int(x) for x in size_str.lower().replace(' ', '').split('x') if x]
        print_long_cm = max(dims) if dims else 0
    except ValueError:
        print_long_cm = 0
    ai_prompt, use_refs = _resolve_preset_ai(preset_key)
    # Web không gửi thông số chi tiết -> tự áp BỘ THÔNG SỐ CHUẨN của preset
    # (vd 'photo': 40 màu, mượt vừa, bỏ mảng <60px) thay vì 0/0 thô.
    from pha.ai_enhance import get_preset as _gp
    _pp = _gp(preset_key)
    if color_limit <= 0:
        color_limit = int(_pp.get('color_limit') or 0)
    api_min_area = int(_pp.get('min_area') or 0)
    api_smooth = int(_pp.get('smooth') or 0)
    # CHÂN DUNG: lấy cờ detail + face_priority TỪ PRESET (giống hệt luồng nhân viên).
    # Web bán hàng chỉ gửi preset 'photo' (ảnh thật) -> face_priority=True -> tự bật
    # YuNet bắt mặt + lượng tử cục bộ chống chảy mặt + tự zoom vào người. Preset khác
    # (vd anime) -> face_priority=False -> KHÔNG đụng (không zoom/refine nhầm).
    api_detail = bool(_pp.get('detail'))
    api_face_priority = bool(_pp.get('face_priority'))

    fss = FileSystemStorage()
    # uuid -> tên DUY NHẤT (tránh 2 upload cùng giây + cùng tên file đè/lẫn nhau);
    # GÁN lại name = fss.save(...) để dùng tên THẬT Django đã ghi (nó tự đổi tên nếu trùng,
    # bỏ giá trị trả về = record trỏ nhầm file job khác -> LẪN ẢNH).
    name = f'{datetime.now():%Y-%m-%d_%H-%M-%S}_api_{uuid.uuid4().hex[:8]}_{upload.name}'
    name = fss.save(name, upload)
    rec = ImageResult.objects.create(
        name=name, status=ImageResult.STATUS_PROCESSING, user='api',
        params={'enhance': enhance, 'preset': preset_key, 'color_limit': color_limit,
                'min_area': api_min_area, 'smooth': api_smooth,
                'detail': api_detail, 'face_priority': api_face_priority,
                'print_size': size_str, 'source': 'thiet-ke'})

    # Chạy NỀN (qua thread như luồng nhân viên) rồi trả id ngay — request chỉ
    # sống ~1-2s nên không dính timeout proxy/nginx/gunicorn nào (AI mất 20-150s).
    _img_executor.submit(process_image, rec.id, name, enhance, None, color_limit,
                         api_min_area, api_smooth, ai_prompt, use_refs, print_long_cm,
                         api_detail, api_face_priority)
    _prune_image_results()
    return _cors(JsonResponse({'ok': True, 'status': 'processing', 'id': rec.id}))


@csrf_exempt
def api_xu_ly_anh_status(request):
    """Tra trạng thái job của /api/xu-ly-anh (id). Cùng khoá API.
    Trả: processing | error | done (kèm URL ảnh tuyệt đối + bảng màu)."""
    from pha.models import AppSetting
    from pha.imageproc import split_list

    def _cors(resp):
        resp['Access-Control-Allow-Origin'] = getattr(settings, 'THIETKE_ALLOW_ORIGIN', '*')
        resp['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        resp['Access-Control-Allow-Headers'] = 'Content-Type, X-API-Key'
        resp['Cache-Control'] = 'no-store'
        return resp

    if request.method == 'OPTIONS':
        return _cors(HttpResponse(status=204))

    want_key = (AppSetting.get('THIETKE_API_KEY', '') or getattr(settings, 'THIETKE_API_KEY', '')).strip()
    got_key = (request.headers.get('X-API-Key') or request.GET.get('key') or '').strip()
    if not want_key or got_key != want_key:
        return _cors(JsonResponse({'ok': False, 'error': 'Sai hoặc thiếu khoá API'}, status=401))

    try:
        rec = ImageResult.objects.get(id=int(request.GET.get('id', 0)))
    except (ImageResult.DoesNotExist, ValueError, TypeError):
        return _cors(JsonResponse({'ok': False, 'status': 'error', 'error': 'Không tìm thấy job.'}, status=404))

    if rec.status == ImageResult.STATUS_PROCESSING:
        from pha.imageproc import mark_if_stuck
        if not mark_if_stuck(rec):          # kẹt quá lâu -> rơi xuống nhánh lỗi
            return _cors(JsonResponse({'ok': True, 'status': 'processing', 'id': rec.id}))
    if rec.status == ImageResult.STATUS_ERROR:
        return _cors(JsonResponse({'ok': True, 'status': 'error', 'id': rec.id,
                                   'error': rec.error_message or 'Xử lý thất bại'}))

    def _u(rel):
        return request.build_absolute_uri('/media/' + rel) if rel else ''
    return _cors(JsonResponse({
        'ok': True, 'status': 'done', 'id': rec.id,
        'img_output': _u(rec.name_output),
        'enhanced': _u(rec.enhanced_name),
        'original': _u(rec.name),
        'colors': split_list(10, rec.colors),
        'warn': rec.error_message or '',
    }))


@csrf_exempt
def api_ketoan(request):
    """API đọc dữ liệu cho phần mềm KẾ TOÁN (ketoan.tranhdali.vn).
    Trả JSON: chi phí sơn theo tháng, tồn kho sơn, số mẻ pha.
    Bảo vệ bằng khoá ?key= (KETOAN_API_KEY). Có CORS để subdomain gọi được.
    """
    origin = getattr(settings, 'KETOAN_ALLOW_ORIGIN', '*')

    def _cors(resp):
        resp['Access-Control-Allow-Origin'] = origin
        resp['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        resp['Access-Control-Allow-Headers'] = 'Content-Type'
        resp['Cache-Control'] = 'no-store'
        return resp

    if request.method == 'OPTIONS':
        return _cors(HttpResponse(status=204))

    key = request.GET.get('key', '')
    if key != _api_key():
        return _cors(JsonResponse({'ok': False, 'error': 'Sai khoá API'}, status=401))

    now = _now()
    this_month = now.strftime('%Y-%m')

    # 12 tháng gần nhất: chi phí sơn + số mẻ
    months, y, m = [], now.year, now.month
    for _ in range(12):
        months.append(f'{y:04d}-{m:02d}')
        m -= 1
        if m == 0:
            m = 12; y -= 1
    months = months[::-1]
    cost_map = dict(ProductionLog.objects.values_list('month').annotate(c=Sum('cost')))
    cnt_map = dict(ProductionLog.objects.values_list('month').annotate(n=Count('id')))
    monthly = [{
        'month': mm, 'label': _fmt_month(mm),
        'paint_cost': round(cost_map.get(mm, 0) or 0),
        'batches': cnt_map.get(mm, 0) or 0,
    } for mm in months]

    # Tồn kho sơn hiện tại
    inventory, total_value = [], 0.0
    for b in mixing.get_bases():
        ps, _ = PaintStock.objects.get_or_create(name=b['name'])
        value = ps.stock / 1000.0 * (ps.price_per_kg or 0)
        total_value += value
        inventory.append({
            'name': b['name'], 'stock_g': round(ps.stock, 1),
            'price_per_kg': round(ps.price_per_kg or 0), 'value': round(value),
        })

    data = {
        'ok': True,
        'source': 'mau.tranhdali.vn',
        'generated_at': now.strftime('%Y-%m-%d %H:%M'),
        'current_month': this_month,
        'summary': {
            'month_paint_cost': round(cost_map.get(this_month, 0) or 0),
            'month_batches': cnt_map.get(this_month, 0) or 0,
            'today_batches': ProductionLog.objects.filter(day=now.strftime('%Y-%m-%d')).count(),
            'inventory_value': round(total_value),
        },
        'monthly': monthly,
        'inventory': inventory,
    }
    return _cors(JsonResponse(data))


@csrf_exempt
def api_luong(request):
    """API LƯƠNG cho phần mềm kế toán (ketoan.tranhdali.vn): mỗi nhân viên/tháng gồm
    CHẤM CÔNG (ngày công, giờ, tăng ca, đi muộn) + SẢN LƯỢNG (tham khảo).
    Lương TÍNH THEO NGÀY CÔNG bên kế toán — API này không tính tiền lương.
    Bảo vệ bằng ?key=KETOAN_API_KEY. Có CORS. ?month=YYYY-MM (mặc định tháng này)."""
    origin = getattr(settings, 'KETOAN_ALLOW_ORIGIN', '*')

    def _cors(resp):
        resp['Access-Control-Allow-Origin'] = origin
        resp['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        resp['Access-Control-Allow-Headers'] = 'Content-Type'
        resp['Cache-Control'] = 'no-store'
        return resp

    if request.method == 'OPTIONS':
        return _cors(HttpResponse(status=204))
    if request.GET.get('key', '') != _api_key():
        return _cors(JsonResponse({'ok': False, 'error': 'Sai khoá API'}, status=401))

    from pha.models import Attendance
    now = _now()
    month = request.GET.get('month') or now.strftime('%Y-%m')

    # Sản lượng (để kế toán tham khảo / tính thưởng nếu cần) — KHÔNG tính tiền ở đây
    _, prows, _ptot = _productivity('month', month)
    pmap = {r['user']: r for r in prows}

    # Chấm công: ngày công + giờ + tăng ca + đi muộn (đầu vào để kế toán tính lương theo NGÀY CÔNG)
    cfg = _att_cfg()
    att = {}
    for r in Attendance.objects.filter(month=month):
        c = _att_calc(r, cfg)
        a = att.setdefault(r.user, {'work_days': 0, 'total_hours': 0.0, 'ot_hours': 0.0,
                                    'ot_pay': 0, 'late_minutes': 0, 'late_fine': 0, 'late_days': 0})
        if r.check_in:
            a['work_days'] += 1
        if c['late_min'] > 0:
            a['late_days'] += 1
        a['total_hours'] += c['hours']; a['ot_hours'] += c['ot_hours']
        a['ot_pay'] += c['ot_pay']; a['late_minutes'] += c['late_min']; a['late_fine'] += c['fine']

    # Nghỉ phép đã duyệt trong tháng (để kế toán trừ/cộng phép)
    from pha.extra_views import _approved_leave_days
    leave_map = _approved_leave_days(month)

    users = sorted(set(pmap) | set(att) | set(leave_map))
    employees = []
    for u in users:
        p = pmap.get(u, {'pha': 0, 'rot_p': 0, 'rot_c': 0, 'sx': 0})
        a = att.get(u, {'work_days': 0, 'total_hours': 0.0, 'ot_hours': 0.0,
                        'ot_pay': 0, 'late_minutes': 0, 'late_fine': 0, 'late_days': 0})
        employees.append({
            'user': u,
            'attendance': {'work_days': a['work_days'], 'total_hours': round(a['total_hours'], 2),
                           'ot_hours': round(a['ot_hours'], 2), 'ot_pay': round(a['ot_pay']),
                           'late_days': a['late_days'], 'late_minutes': a['late_minutes'],
                           'late_fine': round(a['late_fine']),
                           'leave_days': len(leave_map.get(u, ()))},
            'output': {'pha_batches': p['pha'], 'rot_paintings': p['rot_p'],
                       'rot_colors': p['rot_c'], 'sx_paintings': p['sx']},
        })

    data = {
        'ok': True, 'source': 'mau.tranhdali.vn',
        'generated_at': now.strftime('%Y-%m-%d %H:%M'), 'month': month,
        'note': 'Lương tính theo NGÀY CÔNG ở phần mềm kế toán. API này chỉ cung cấp '
                'số ngày công, giờ, tăng ca, đi muộn (+ sản lượng để tham khảo). '
                'Không tính tiền lương ở đây.',
        'schedule': {'work_start': cfg['start'], 'work_end': cfg['end'],
                     'work_days': sorted(cfg['workdays']),
                     'work_days_label': [_WEEKDAYS[i] for i in sorted(cfg['workdays'])],
                     'late_grace_min': cfg['grace'], 'late_fine_per_min': cfg['fine_min'],
                     'late_fine_fixed': cfg['fine_fixed'], 'ot_rate_per_hour': cfg['ot_rate']},
        'employees': employees,
        'totals': {
            'work_days': sum(e['attendance']['work_days'] for e in employees),
            'ot_hours': round(sum(e['attendance']['ot_hours'] for e in employees), 2),
            'ot_pay': sum(e['attendance']['ot_pay'] for e in employees),
            'late_fine': sum(e['attendance']['late_fine'] for e in employees),
        },
    }
    return _cors(JsonResponse(data))


@csrf_exempt
@staff_required
def kho_son(request):
    """Quản lý tồn kho màu sơn gốc (chỉ chủ)."""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        action = request.POST.get('action')
        try:
            val = float(request.POST.get('value') or 0)
        except ValueError:
            val = 0
        ps, _ = PaintStock.objects.get_or_create(name=name)
        if action == 'add_stock':
            ps.stock = round(ps.stock + val, 1); ps.save()
            messages.info(request, f'Đã nhập thêm {val:g}g vào "{name}".')
        elif action == 'set_stock':
            ps.stock = val; ps.save()
            messages.info(request, f'Đã đặt tồn kho "{name}" = {val:g}g.')
        elif action == 'set_threshold':
            ps.low_threshold = val; ps.save()
            messages.info(request, f'Đã đặt ngưỡng cảnh báo "{name}" = {val:g}g.')
        elif action == 'set_price':
            ps.price_per_kg = val; ps.save()
            messages.info(request, f'Đã đặt giá "{name}" = {val:,.0f} đ/kg.')
        return redirect('/kho-son')

    items = []
    total_value = 0
    for b in mixing.get_bases():
        ps, _ = PaintStock.objects.get_or_create(name=b['name'])
        value = ps.stock * ps.price_per_kg / 1000.0
        total_value += value
        items.append({
            'name': b['name'], 'rgb': b['rgb'], 'stock': round(ps.stock, 1),
            'threshold': ps.low_threshold, 'price': ps.price_per_kg,
            'value': round(value),
            'low': ps.low_threshold > 0 and ps.stock <= ps.low_threshold,
        })
    return render(request, 'kho_son.html', {'items': items, 'total_value': round(total_value),
                                            'low_stock': _low_stock_names()})


def _now():
    return datetime.now(_VN) if _VN else datetime.now()


def _aggregate(qs):
    acc = {}
    for log in qs:
        for c in (log.components or []):
            try:
                acc[c['name']] = round(acc.get(c['name'], 0) + float(c['grams']), 2)
            except (KeyError, TypeError, ValueError):
                continue
    return sorted([{'name': k, 'grams': v} for k, v in acc.items()], key=lambda x: -x['grams'])


def _fmt_month(m):
    try:
        return datetime.strptime(m, '%Y-%m').strftime('%m/%Y')
    except ValueError:
        return m


def _stats_qs(range_, month_param):
    """Trả (label, queryset ProductionLog) theo khoảng: today / week / month."""
    now = _now()
    if range_ == 'week':
        d = now.date()
        monday = d - timedelta(days=d.weekday())
        sunday = monday + timedelta(days=6)
        qs = ProductionLog.objects.filter(day__gte=monday.strftime('%Y-%m-%d'),
                                          day__lte=sunday.strftime('%Y-%m-%d'))
        label = f"Tuần này ({monday.strftime('%d/%m')} – {sunday.strftime('%d/%m/%Y')})"
    elif range_ == 'month':
        m = month_param or now.strftime('%Y-%m')
        qs = ProductionLog.objects.filter(month=m)
        label = "Tháng " + _fmt_month(m)
    else:
        qs = ProductionLog.objects.filter(day=now.strftime('%Y-%m-%d'))
        label = "Hôm nay (" + now.strftime('%d/%m/%Y') + ")"
    return label, qs


def _stats(range_, month_param):
    label, qs = _stats_qs(range_, month_param)
    return label, _aggregate(qs)


@csrf_exempt
@staff_required
def cong_thuc_mau(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_base':
            ok, msg = mixing.add_base(request.POST.get('name', ''), request.POST.get('hex', ''))
            messages.info(request, msg)
            return redirect('cong_thuc_mau')
        if action == 'delete_base':
            n = mixing.delete_base(request.POST.get('name', ''))
            messages.info(request, "Đã xoá màu gốc." if n else "Không tìm thấy.")
            return redirect('cong_thuc_mau')
        if action == 'save_recipe':
            names = request.POST.getlist('base_name')
            weights = request.POST.getlist('base_weight')
            components = [{'name': n, 'grams': w} for n, w in zip(names, weights)]
            ok, msg = recipes.add_recipe(request.POST.get('dali', ''), request.POST.get('hex', ''), components)
            messages.info(request, msg)
            return redirect('cong_thuc_mau')
        if action == 'delete_recipe':
            recipes.delete_recipe(request.POST.get('dali', ''))
            messages.info(request, "Đã xoá công thức.")
            return redirect('cong_thuc_mau')

    rec_list = []
    for r in recipes.get_all():
        rec_list.append({
            'dali': r['dali'], 'hex': r['hex'], 'components': r['components'],
            'total': recipes.total_grams(r), 'formula': recipes.as_formula(r),
        })
    months = sorted(set(ProductionLog.objects.values_list('month', flat=True)), reverse=True)
    stat_months = [{'value': m, 'label': _fmt_month(m)} for m in months]
    label, rows = _stats('today', None)
    return render(request, 'cong_thuc_mau.html', {
        'bases': mixing.get_bases(), 'recipes': rec_list,
        'stat_months': stat_months, 'stat_label': label, 'stat_rows': rows,
        'low_stock': _low_stock_names(),
    })


@csrf_exempt
@staff_required
def thong_ke(request):
    label, rows = _stats(request.GET.get('range', 'today'), request.GET.get('month'))
    return JsonResponse({'label': label, 'rows': rows})


@csrf_exempt
@login_required(login_url='/login')
def lich_su(request):
    """Lịch sử các mẻ đã pha (mới nhất trước). ?range=today|all."""
    now = _now()
    if request.GET.get('range') == 'today':
        qs = ProductionLog.objects.filter(day=now.strftime('%Y-%m-%d'))
    else:
        qs = ProductionLog.objects.all()
    rows = []
    for log in qs.order_by('-created_time')[:100]:
        t = log.created_time
        try:
            t = t.astimezone(_VN) if _VN else t
        except Exception:
            pass
        rows.append({'dt': t.strftime('%d/%m %H:%M'), 'dali': log.dali,
                     'mult': '×' + ('%g' % log.multiplier), 'user': log.user or ''})
    return JsonResponse({'rows': rows})


@csrf_exempt
def export_thong_ke_excel(request):
    """Xuất báo cáo lượng màu đã pha ra Excel (.xlsx) theo khoảng thời gian."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    range_ = request.GET.get('range', 'month')
    month = request.GET.get('month')
    label, qs = _stats_qs(range_, month)
    rows = _aggregate(qs)

    wb = Workbook()
    head_fill = PatternFill('solid', fgColor='2E7D32')
    head_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')

    # Sheet 1: tổng theo màu gốc
    ws = wb.active
    ws.title = "Tong theo mau"
    ws.append(["BÁO CÁO LƯỢNG MÀU ĐÃ PHA"])
    ws.append([label])
    ws.append(["Màu gốc", "Tổng đã dùng (g)"])
    for c in ws[3]:
        c.fill = head_fill; c.font = head_font; c.alignment = center
    for u in rows:
        ws.append([u['name'], u['grams']])
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18

    # Sheet 2: chi tiết từng mẻ pha
    ws2 = wb.create_sheet("Chi tiet pha")
    ws2.append(["Ngày giờ", "Mã DALI", "Hệ số nhân", "Tổng (g)", "Chi tiết"])
    for c in ws2[1]:
        c.fill = head_fill; c.font = head_font; c.alignment = center
    for log in qs.order_by('created_time'):
        detail = " + ".join(f"{x.get('name')} {x.get('grams')}g" for x in (log.components or []))
        t = log.created_time
        try:
            t = t.astimezone(_VN) if _VN else t
        except Exception:
            pass
        ws2.append([t.strftime('%d/%m/%Y %H:%M'), log.dali, f"x{('%g' % log.multiplier)}",
                    log.total, detail])
    for col, w in zip('ABCDE', (18, 14, 12, 12, 60)):
        ws2.column_dimensions[col].width = w

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="bao_cao_mau.xlsx"'
    wb.save(resp)
    return resp


@csrf_exempt
@login_required(login_url='/login')
def pha(request):
    if request.method != 'POST':
        return HttpResponseNotFound("POST only")
    dali = request.POST.get('dali', '').strip()
    try:
        mult = float(request.POST.get('multiplier') or 1)
    except ValueError:
        mult = 1.0
    if mult <= 0:
        mult = 1.0
    rec = next((r for r in recipes.get_all() if r['dali'].strip().lower() == dali.lower()), None)
    if not rec:
        return JsonResponse({'ok': False, 'msg': 'Không tìm thấy công thức.'})
    comps = [{'name': c['name'], 'grams': round(c['grams'] * mult, 2)} for c in rec['components']]
    total = round(sum(c['grams'] for c in comps), 2)
    # Chi phí sơn của mẻ (theo giá hiện tại, đồng)
    prices = {p.name: p.price_per_kg for p in PaintStock.objects.all()}
    cost = round(sum(c['grams'] * prices.get(c['name'], 0) / 1000.0 for c in comps))
    now = _now()
    ProductionLog.objects.create(
        day=now.strftime('%Y-%m-%d'), month=now.strftime('%Y-%m'),
        dali=rec['dali'], hex=rec['hex'], multiplier=mult, components=comps, total=total,
        user=request.user.username, cost=cost,
    )
    # Trừ kho sơn theo lượng đã dùng (chỉ trừ màu có theo dõi tồn kho)
    for c in comps:
        PaintStock.objects.filter(name=c['name']).update(stock=F('stock') - c['grams'])
    return JsonResponse({'ok': True, 'msg': f'Đã ghi nhận pha {rec["dali"]} ×{("%g" % mult)}'})


@csrf_exempt
@login_required(login_url='/login')
def mobile(request):
    bases = {b['name']: b['rgb'] for b in mixing.get_bases()}
    rec_list = []
    for r in recipes.get_all():
        comps = [{'name': c['name'], 'grams': c['grams'], 'rgb': bases.get(c['name'])}
                 for c in r['components']]
        rec_list.append({'dali': r['dali'], 'hex': r['hex'], 'components': comps,
                         'total': recipes.total_grams(r)})
    return render(request, 'mobile.html', {'recipes': rec_list})


@csrf_exempt
@staff_required
def quan_ly(request):
    """App ĐIỆN THOẠI cho quản lý: nhập kho sơn + xem nhanh dashboard."""
    now = _now()
    today_n = ProductionLog.objects.filter(day=now.strftime('%Y-%m-%d')).count()
    month_n = ProductionLog.objects.filter(month=now.strftime('%Y-%m')).count()
    month_cost = ProductionLog.objects.filter(month=now.strftime('%Y-%m')).aggregate(s=Sum('cost'))['s'] or 0
    since = (now.date() - timedelta(days=29)).strftime('%Y-%m-%d')
    usage30 = defaultdict(float)
    for log in ProductionLog.objects.filter(day__gte=since):
        for c in (log.components or []):
            try:
                usage30[c['name']] += float(c['grams'])
            except (KeyError, TypeError, ValueError):
                pass
    items, need_buy = [], []
    total_value = 0.0
    for b in mixing.get_bases():
        ps, _ = PaintStock.objects.get_or_create(name=b['name'])
        low = ps.low_threshold > 0 and ps.stock <= ps.low_threshold
        value = ps.stock / 1000.0 * (ps.price_per_kg or 0)
        total_value += value
        # Ước tính số ngày còn dùng được theo mức dùng 30 ngày gần nhất
        avg_day = usage30.get(b['name'], 0) / 30.0
        days = int(ps.stock / avg_day) if avg_day > 0 else None
        if low or (days is not None and days < 7):
            level, bar = 'low', (min(100, round(days / 30.0 * 100)) if days is not None else 8)
        elif days is not None and days < 14:
            level, bar = 'warn', min(100, round(days / 30.0 * 100))
        else:
            level, bar = 'ok', (min(100, round(days / 30.0 * 100)) if days is not None else 100)
        items.append({
            'name': b['name'], 'rgb': b['rgb'], 'stock': round(ps.stock, 1), 'low': low,
            'value': f'{round(value):,.0f}'.replace(',', '.'), 'days': days,
            'level': level, 'bar': bar, 'price': round(ps.price_per_kg or 0),
        })
        suggest = max(0, round(usage30.get(b['name'], 0) - ps.stock))
        if suggest > 0:
            need_buy.append({'name': b['name'], 'suggest': suggest})
    from pha.models import Painting
    return render(request, 'quan_ly.html', {
        'items': items, 'today_n': today_n, 'month_n': month_n, 'month_cost': round(month_cost),
        'low_stock': _low_stock_names(), 'need_buy': need_buy, 'today_label': now.strftime('%d/%m/%Y'),
        'total_value': f'{round(total_value):,.0f}'.replace(',', '.'),
        'paintings_json': json.dumps([_painting_dict(p) for p in Painting.objects.all()]),
        'staff_users': _staff_users(),
        'paint_sizes': _paint_sizes(),
    })


@csrf_exempt
@staff_required
def quan_ly_nhap(request):
    """AJAX: nhập thêm / đặt lại tồn kho 1 màu. Trả JSON."""
    if request.method != 'POST':
        return HttpResponseNotFound('POST only')
    name = request.POST.get('name', '').strip()
    action = request.POST.get('action')
    try:
        val = float(request.POST.get('value') or 0)
    except ValueError:
        val = 0
    ps, _ = PaintStock.objects.get_or_create(name=name)
    if action == 'add':
        ps.stock = round(ps.stock + val, 1)
    elif action == 'set':
        ps.stock = val
    else:
        return JsonResponse({'ok': False})
    ps.save()
    low = ps.low_threshold > 0 and ps.stock <= ps.low_threshold
    return JsonResponse({'ok': True, 'name': name, 'stock': round(ps.stock, 1), 'low': low})


def manifest_ql(request):
    data = {
        "name": "Quản lý kho sơn", "short_name": "Quản lý",
        "start_url": "/quan-ly", "scope": "/", "display": "standalone", "orientation": "portrait",
        "background_color": "#ffffff", "theme_color": "#0d6efd",
        "icons": [
            {"src": "/media/icon-ql-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": "/media/icon-ql-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
        ],
    }
    return JsonResponse(data, content_type='application/manifest+json')


def manifest(request):
    data = {
        "name": "Công thức pha DALI", "short_name": "Pha màu",
        "start_url": "/app", "scope": "/", "display": "standalone", "orientation": "portrait",
        "background_color": "#ffffff", "theme_color": "#2E7D32",
        "icons": [
            {"src": "/media/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": "/media/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
        ],
    }
    return JsonResponse(data, content_type='application/manifest+json')


def manifest_app(request):
    """Manifest cho APP THỐNG NHẤT (cài màn hình chính, mở /home)."""
    data = {
        "name": "DALI Tranh số hoá", "short_name": "DALI",
        "start_url": "/home", "scope": "/", "display": "standalone", "orientation": "portrait",
        "background_color": "#ffffff", "theme_color": "#2E7D32",
        "icons": [
            {"src": "/media/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any"},
            {"src": "/media/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": "/media/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
        ],
    }
    return JsonResponse(data, content_type='application/manifest+json')


@csrf_exempt
@login_required(login_url='/login')
def home(request):
    """Màn hình APP thống nhất (lưới chức năng theo vai trò). Cài được vào màn hình chính."""
    if request.user.is_staff:
        tiles = [
            {'url': '/', 'icon': 'bi-droplet-half', 'label': 'Công thức pha', 'bg': '#2E7D32'},
            {'url': '/ma-tranh', 'icon': 'bi-paint-bucket', 'label': 'Rót màu', 'bg': '#6F42C1'},
            {'url': '/san-xuat', 'icon': 'bi-box2-heart', 'label': 'Sản xuất', 'bg': '#198754'},
            {'url': '/nang-suat', 'icon': 'bi-people-fill', 'label': 'Năng suất', 'bg': '#0D6EFD'},
            {'url': '/loi-nhuan', 'icon': 'bi-graph-up-arrow', 'label': 'Lợi nhuận', 'bg': '#FD7E14'},
            {'url': '/cham-cong-quan-ly', 'icon': 'bi-fingerprint', 'label': 'Chấm công', 'bg': '#20C997'},
            {'url': '/kho-son', 'icon': 'bi-box-seam', 'label': 'Kho sơn', 'bg': '#8D6E63'},
            {'url': '/xu-ly-anh', 'icon': 'bi-image', 'label': 'Xử lý ảnh', 'bg': '#E83E8C'},
            {'url': '/dashboard', 'icon': 'bi-speedometer2', 'label': 'Dashboard', 'bg': '#343A40'},
            {'url': '/nhan-vien', 'icon': 'bi-people', 'label': 'Nhân viên', 'bg': '#6C757D'},
            {'url': '/quan-ly', 'icon': 'bi-phone-vibrate', 'label': 'App kho (QL)', 'bg': '#0DCAF0'},
            {'url': '/cham-cong', 'icon': 'bi-clock', 'label': 'Chấm công của tôi', 'bg': '#0F9D58'},
        ]
    else:
        tiles = [
            {'url': '/app', 'icon': 'bi-droplet-half', 'label': 'Pha màu & Rót màu', 'bg': '#2E7D32'},
            {'url': '/cham-cong', 'icon': 'bi-fingerprint', 'label': 'Chấm công', 'bg': '#198754'},
        ]
    return render(request, 'home.html', {'tiles': tiles})


def service_worker(request):
    js = (
        "const CACHE='pha-v2';\n"
        "self.addEventListener('install', function(e){ self.skipWaiting(); });\n"
        "self.addEventListener('activate', function(e){ e.waitUntil(self.clients.claim()); });\n"
        "self.addEventListener('fetch', function(e){\n"
        "  if(e.request.method!=='GET') return;\n"
        "  e.respondWith(fetch(e.request).then(function(r){\n"
        "    try{ var c=r.clone(); caches.open(CACHE).then(function(ch){ ch.put(e.request, c); }); }catch(_){}\n"
        "    return r;\n"
        "  }).catch(function(){ return caches.match(e.request); }));\n"
        "});\n"
        "self.addEventListener('push', function(e){\n"
        "  var d={}; try{ d=e.data.json(); }catch(_){ try{ d={body:e.data.text()}; }catch(__){ d={}; } }\n"
        "  var title=d.title||'🎨 Mã màu cần rót';\n"
        "  var opts={ body:d.body||'', icon:d.icon||'/media/icon-192.png', badge:'/media/icon-192.png',\n"
        "    tag:d.tag||'rot', renotify:true, vibrate:[120,60,120], data:{url:d.url||'/app'} };\n"
        "  e.waitUntil(self.registration.showNotification(title, opts));\n"
        "});\n"
        "self.addEventListener('notificationclick', function(e){\n"
        "  e.notification.close();\n"
        "  var url=(e.notification.data&&e.notification.data.url)||'/app';\n"
        "  e.waitUntil(clients.matchAll({type:'window',includeUncontrolled:true}).then(function(cl){\n"
        "    for(var i=0;i<cl.length;i++){ var c=cl[i]; if('focus' in c){ try{ c.navigate&&c.navigate(url); }catch(_){}; return c.focus(); } }\n"
        "    if(clients.openWindow) return clients.openWindow(url);\n"
        "  }));\n"
        "});\n"
    )
    return HttpResponse(js, content_type='application/javascript')


@csrf_exempt
@staff_required
def dali_colors(request):
    """Bảng màu DALI: xem / tìm / thêm-sửa / xoá / nạp lại (chỉ chủ)."""
    from pha import dali_match
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            ok, msg = dali_match.add_entry(request.POST.get('hex', ''), request.POST.get('dali', ''))
            messages.info(request, msg)
        elif action == 'delete':
            n = dali_match.delete_entry(request.POST.get('hex', ''), request.POST.get('dali') or None)
            messages.info(request, f'Đã xoá {n} mục.' if n else 'Không tìm thấy mục để xoá.')
        elif action == 'reload':
            n = dali_match.reload_reference()
            messages.info(request, f'Đã nạp lại {n} màu từ file.')
        return redirect('/dali-colors')

    query = (request.GET.get('q') or '').strip().lower()
    recent = request.GET.get('recent') == '1'
    items = dali_match.get_recent() if recent else dali_match.get_all()
    if query:
        items = [it for it in items if query in it['hex'].lower() or query in it['dali'].lower()]
    total = dali_match.reference_size()
    shown = items[:500]
    return render(request, 'dali_colors.html', {
        'items': shown, 'total': total, 'query': request.GET.get('q') or '',
        'found': len(items), 'truncated': len(items) > 500, 'recent': recent,
    })


# ===================== RÓT MÀU (theo mã tranh) =====================
def _detect_color_count(path):
    """Lấy SỐ MÀU của ảnh mẫu. Ưu tiên nhờ AI đọc bảng chú giải (chính xác); nếu
    chưa có khoá AI / lỗi thì ước lượng bằng pixel (gần đúng). Số GỢI Ý — sửa được."""
    try:
        from pha.ai_enhance import ai_count_colors, is_configured
        if is_configured():
            n = ai_count_colors(path)
            if n:
                return n
    except Exception:
        pass
    return _detect_color_count_pixel(path)


def _detect_color_count_pixel(path):
    """Ước lượng số màu bằng pixel: đếm các mảng màu phẳng lớn, gộp màu gần giống."""
    try:
        from PIL import Image
        import numpy as np
        img = Image.open(path).convert('RGB')
        img.thumbnail((400, 400))                      # nhỏ lại cho nhanh
        arr = np.asarray(img).reshape(-1, 3).astype(int)
        q = (arr // 24) * 24                            # gom bớt sai khác do khử răng cưa
        colors, counts = np.unique(q, axis=0, return_counts=True)
        total = counts.sum()
        keep = counts >= total * 0.012                 # chỉ giữ màu chiếm >=1.2% diện tích
        colors, counts = colors[keep], counts[keep]
        order = np.argsort(-counts)                     # nhiều diện tích trước
        reps = []
        for c in colors[order]:
            if all(int(np.sum((c - r) ** 2)) > 42 * 42 for r in reps):  # cách nhau đủ xa
                reps.append(c)
        return len(reps)
    except Exception:
        return 0


def _painting_map():
    """{code(lower): Painting} để tra nhanh mã tranh trong catalog."""
    from pha.models import Painting
    return {p.code.strip().lower(): p for p in Painting.objects.all()}


def _remove_media(name):
    """Xoá 1 file trong MEDIA_ROOT (bỏ qua nếu không có)."""
    if not name:
        return
    try:
        os.remove(os.path.join(settings.MEDIA_ROOT, name))
    except OSError:
        pass


def _painting_dict(p):
    """Mã tranh dạng dict cho JSON (mã + số màu + URL ảnh)."""
    return {'code': p.code, 'count': p.color_count,
            'image': ('/media/' + p.image) if p.image else ''}


def _painting_count(code):
    """Trả (số_màu, Painting|None) theo mã tranh."""
    from pha.models import Painting
    p = Painting.objects.filter(code__iexact=(code or '').strip()).first()
    return (p.color_count if p else 0), p


def _paint_sizes():
    """Danh sách kích thước tranh (gợi ý). Chủ có thể đặt AppSetting 'PAINT_SIZES'
    (cách nhau dấu phẩy) để đổi; mặc định các khổ phổ biến."""
    from pha.models import AppSetting
    raw = (AppSetting.get('PAINT_SIZES', '') or '').strip()
    if raw:
        out = [s.strip() for s in raw.split(',') if s.strip()]
        if out:
            return out
    return ['20x20', '30x30', '30x37.5']


def _norm_size(s):
    """Chuẩn hoá kích thước: bỏ khoảng trắng, thường hoá 'x' (40 X 50 -> 40x50)."""
    return (s or '').strip().replace(' ', '').lower()


# Cửa sổ CHỐNG GHI TRÙNG (giây): cùng người + cùng mã trong khoảng này -> coi là bấm 2
# lần / gửi lại / quét trùng -> hỏi lại trước khi ghi (force=1 để ghi tiếp khi cố ý).
POUR_DEDUP_SECONDS = 45


def _record_pour(painting, qty, color_count, user, req=None, size=''):
    """Ghi 1 lượt rót màu vào nhật ký; nếu có yêu cầu (req) thì đánh dấu đã rót."""
    from pha.models import PourLog, PourRequest
    now = _now()
    qty = max(1, int(qty or 1))
    PourLog.objects.create(
        day=now.strftime('%Y-%m-%d'), month=now.strftime('%Y-%m'),
        painting=painting, size=_norm_size(size), colors=[],
        color_count=int(color_count or 0), qty=qty,
        user=user, request_id=(req.id if req else None),
    )
    if req and req.status != PourRequest.STATUS_DONE:
        req.status = PourRequest.STATUS_DONE
        req.done_by = user
        req.done_time = now
        req.save(update_fields=['status', 'done_by', 'done_time'])


def _pour_stats_qs(range_, month_param):
    """Trả (label, queryset PourLog) theo khoảng: today / week / month."""
    from pha.models import PourLog
    now = _now()
    if range_ == 'week':
        d = now.date()
        monday = d - timedelta(days=d.weekday())
        sunday = monday + timedelta(days=6)
        qs = PourLog.objects.filter(day__gte=monday.strftime('%Y-%m-%d'),
                                    day__lte=sunday.strftime('%Y-%m-%d'))
        label = f"Tuần này ({monday.strftime('%d/%m')} – {sunday.strftime('%d/%m/%Y')})"
    elif range_ == 'month':
        m = month_param or now.strftime('%Y-%m')
        qs = PourLog.objects.filter(month=m)
        label = "Tháng " + _fmt_month(m)
    else:
        qs = PourLog.objects.filter(day=now.strftime('%Y-%m-%d'))
        label = "Hôm nay (" + now.strftime('%d/%m/%Y') + ")"
    return label, qs


def _pour_aggregate(qs):
    """Tổng hợp: số lượt rót, tổng số tranh, tổng số màu đã rót; chi tiết theo mã tranh
    (cho Excel) và theo NGÀY (cho biểu đồ cột)."""
    pours = 0
    paintings = 0
    colors_total = 0
    acc = {}      # code -> {'painting','pours','qty','cc','colors'}
    dayacc = {}   # YYYY-MM-DD -> {'paintings','colors','pours'}
    sizeacc = {}  # size -> {'paintings','pours'}
    for log in qs:
        pours += 1
        q = max(1, int(log.qty or 1))
        cc = int(log.color_count or 0)
        paintings += q
        colors_total += cc * q
        row = acc.setdefault(log.painting, {'painting': log.painting, 'pours': 0,
                                            'qty': 0, 'cc': cc, 'colors': 0})
        row['pours'] += 1
        row['qty'] += q
        row['colors'] += cc * q
        row['cc'] = cc
        d = dayacc.setdefault(log.day, {'paintings': 0, 'colors': 0, 'pours': 0})
        d['paintings'] += q
        d['colors'] += cc * q
        d['pours'] += 1
        sz = (log.size or '').strip() or '(chưa ghi)'
        s = sizeacc.setdefault(sz, {'paintings': 0, 'pours': 0})
        s['paintings'] += q
        s['pours'] += 1
    rows = sorted(acc.values(), key=lambda x: -x['qty'])
    daily = []
    for k in sorted(dayacc.keys()):
        dd = dayacc[k]
        try:
            lbl = datetime.strptime(k, '%Y-%m-%d').strftime('%d/%m')
        except ValueError:
            lbl = k
        daily.append({'label': lbl, 'paintings': dd['paintings'],
                      'colors': dd['colors'], 'pours': dd['pours']})
    by_size = sorted([{'size': k, 'paintings': v['paintings'], 'pours': v['pours']}
                      for k, v in sizeacc.items()], key=lambda x: -x['paintings'])
    return {'pours': pours, 'paintings': paintings, 'colors_total': colors_total,
            'rows': rows, 'daily': daily, 'by_size': by_size}


def _staff_users():
    """Danh sách tài khoản có thể giao việc (không gồm superuser)."""
    return list(User.objects.filter(is_superuser=False)
                .order_by('is_staff', 'username').values_list('username', flat=True))


@csrf_exempt
@staff_required
def ma_tranh(request):
    """Danh mục MÃ TRANH + giao việc rót màu + thống kê + lịch sử (cho quản lý)."""
    from pha.models import Painting, PourRequest, PourLog

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'save_painting':
            code = (request.POST.get('code') or '').strip()
            if not code:
                messages.error(request, 'Thiếu mã tranh.')
            else:
                obj = Painting.objects.filter(code__iexact=code).first()
                image_name = obj.image if obj else ''
                up = request.FILES.get('image')
                new_upload = False
                if up and up.content_type and up.content_type.startswith('image/'):
                    fss = FileSystemStorage()
                    new_name = fss.save(f'painting_{datetime.now():%Y-%m-%d_%H-%M-%S}_{up.name}', up)
                    _remove_media(image_name)        # xoá ảnh cũ (nếu thay)
                    image_name = new_name
                    new_upload = True
                elif request.POST.get('remove_image') == '1':
                    _remove_media(image_name)
                    image_name = ''
                # Số màu: ưu tiên số người nhập; nếu để trống thì tự đếm từ ảnh mới
                raw_cc = (request.POST.get('color_count') or '').strip()
                if raw_cc.isdigit():
                    color_count = int(raw_cc)
                elif new_upload:
                    color_count = _detect_color_count(os.path.join(settings.MEDIA_ROOT, image_name))
                else:
                    color_count = obj.color_count if obj else 0
                if obj:
                    obj.code, obj.image, obj.color_count = code, image_name, color_count
                    obj.save()
                    messages.info(request, f'Đã cập nhật mã tranh {code} ({color_count} màu).')
                else:
                    Painting.objects.create(code=code, image=image_name, color_count=color_count)
                    messages.info(request, f'Đã lưu mã tranh {code} ({color_count} màu).')
        elif action == 'delete_painting':
            dp = Painting.objects.filter(code__iexact=(request.POST.get('code') or '').strip()).first()
            if dp:
                _remove_media(dp.image)
                dp.delete()
                messages.info(request, 'Đã xoá mã tranh.')
            else:
                messages.info(request, 'Không tìm thấy.')
        elif action == 'add_request':
            code = (request.POST.get('painting') or '').strip()
            p = Painting.objects.filter(code__iexact=code).first()
            if not p:
                messages.error(request, f'Mã tranh "{code}" chưa có trong danh mục.')
            else:
                try:
                    qty = max(1, int(request.POST.get('qty') or 1))
                except ValueError:
                    qty = 1
                req = PourRequest.objects.create(
                    painting=p.code, size=_norm_size(request.POST.get('size')),
                    colors=[], qty=qty,
                    note=(request.POST.get('note') or '').strip(),
                    assignee=(request.POST.get('assignee') or '').strip(),
                    created_by=request.user.username,
                )
                from pha import push
                push.notify_pour(req)
                messages.info(request, f'Đã giao rót {p.code} ×{qty}'
                              + (f' ({req.size})' if req.size else '') + '.')
        elif action == 'delete_request':
            PourRequest.objects.filter(id=request.POST.get('id')).delete()
            messages.info(request, 'Đã xoá yêu cầu.')
        elif action == 'done_request':
            req = PourRequest.objects.filter(id=request.POST.get('id')).first()
            if req and req.status != PourRequest.STATUS_DONE:
                cc, _ = _painting_count(req.painting)
                _record_pour(req.painting, req.qty, cc, request.user.username, req, size=req.size)
                messages.info(request, f'Đã đánh dấu rót xong {req.painting}.')
        return redirect('/ma-tranh')

    paintings = list(Painting.objects.all())
    # Mã đã số hoá + lưu vào kho mã (ảnh = kho_ma/<id>_design.png) -> nút "Sửa màu"
    # mở lại file cũ trong /xu-ly-anh để chỉnh HEX/mã DALI, khỏi chạy lại AI.
    import re as _re
    for p in paintings:
        m = _re.match(r'kho_ma/(m-[0-9a-f]+)_', p.image or '')
        p.reopen_id = m.group(1) if m else ''
    pmap = {p.code.strip().lower(): p for p in paintings}
    pending = list(PourRequest.objects.filter(status=PourRequest.STATUS_PENDING))
    for r in pending:
        pp = pmap.get(r.painting.strip().lower())
        r.image = ('/media/' + pp.image) if (pp and pp.image) else ''
        r.count = pp.color_count if pp else 0
    done = list(PourRequest.objects.filter(status=PourRequest.STATUS_DONE)
                .order_by('-done_time', '-id')[:30])
    for r in done:
        try:
            r.done_disp = (r.done_time.astimezone(_VN) if _VN and r.done_time else r.done_time)
        except Exception:
            r.done_disp = r.done_time
    months = sorted(set(PourLog.objects.values_list('month', flat=True)), reverse=True)
    stat_months = [{'value': m, 'label': _fmt_month(m)} for m in months]
    label, agg = _pour_stats('today', None)
    return render(request, 'ma_tranh.html', {
        'paintings': paintings,
        'paintings_json': json.dumps([_painting_dict(p) for p in paintings]),
        'pending': pending, 'done': done,
        'staff_users': _staff_users(),
        'stat_months': stat_months, 'stat_label': label, 'stat_agg': agg,
        'paint_sizes': _paint_sizes(),
        'low_stock': _low_stock_names(),
    })


def _pour_stats(range_, month_param):
    label, qs = _pour_stats_qs(range_, month_param)
    return label, _pour_aggregate(qs)


@csrf_exempt
@login_required(login_url='/login')
def rot_mau_app(request):
    """Đã gộp vào app nhân viên /app (pha + rót chung 1 app) -> chuyển hướng."""
    return redirect('/app')


@csrf_exempt
@login_required(login_url='/login')
def rot(request):
    """Ghi nhận đã rót xong 1 mã tranh. POST: painting, qty, [request_id], [colors_json]."""
    if request.method != 'POST':
        return HttpResponseNotFound('POST only')
    from pha.models import PourRequest
    req = None
    rid = request.POST.get('request_id')
    if rid:
        req = PourRequest.objects.filter(id=rid).first()
        if not req:
            return JsonResponse({'ok': False, 'msg': 'Yêu cầu không tồn tại.'})
        if req.status == PourRequest.STATUS_DONE:
            return JsonResponse({'ok': False, 'msg': 'Yêu cầu đã được rót trước đó.'})
        painting, qty, size = req.painting, req.qty, req.size
    else:
        painting = (request.POST.get('painting') or '').strip()
        if not painting:
            return JsonResponse({'ok': False, 'msg': 'Thiếu mã tranh.'})
        try:
            qty = max(1, int(request.POST.get('qty') or 1))
        except ValueError:
            qty = 1
        size = (request.POST.get('size') or '').strip()
    # CHỐNG GHI TRÙNG: cùng người vừa ghi đúng mã này trong POUR_DEDUP_SECONDS giây ->
    # nhiều khả năng bấm 2 lần / gửi lại / quét trùng. Trả 'dup' để app HỎI LẠI; chỉ ghi
    # tiếp khi gửi force=1 (người dùng xác nhận cố ý rót thêm). (Rót lại CÙNG yêu cầu đã
    # bị chặn riêng bằng trạng thái DONE ở trên.)
    if request.POST.get('force') != '1':
        from pha.models import PourLog
        from django.utils import timezone
        from datetime import timedelta
        recent = (PourLog.objects
                  .filter(painting=painting, user=request.user.username,
                          created_time__gte=timezone.now() - timedelta(seconds=POUR_DEDUP_SECONDS))
                  .order_by('-created_time').first())
        if recent:
            secs = max(1, int((timezone.now() - recent.created_time).total_seconds()))
            return JsonResponse({'ok': False, 'dup': True, 'painting': painting,
                                 'msg': f'Mã {painting} vừa được ghi {secs}s trước. '
                                        f'Bấm xác nhận nếu CHẮC CHẮN rót thêm lần nữa.'})
    cc, _ = _painting_count(painting)
    _record_pour(painting, qty, cc, request.user.username, req, size=size)
    return JsonResponse({'ok': True, 'msg': f'Đã ghi rót {painting} ×{qty}'})


@csrf_exempt
@staff_required
def cap_nhat_so_mau(request):
    """Tính lại SỐ MÀU tự động từ ẢNH MẪU (AI đọc bảng chú giải, fallback pixel) và
    cập nhật vào mã tranh. action=one (1 mã) hoặc all_zero (các mã đang để 0/thiếu)."""
    if request.method != 'POST':
        return HttpResponseNotFound('POST only')
    from pha.models import Painting
    action = request.POST.get('action') or 'one'
    if action == 'one':
        code = (request.POST.get('code') or '').strip()
        p = Painting.objects.filter(code__iexact=code).first()
        if not p:
            return JsonResponse({'ok': False, 'msg': 'Không tìm thấy mã tranh.'})
        if not p.image:
            return JsonResponse({'ok': False, 'msg': f'{p.code} chưa có ảnh mẫu để đếm.'})
        n = _detect_color_count(os.path.join(settings.MEDIA_ROOT, p.image))
        if n:
            p.color_count = n
            p.save(update_fields=['color_count'])
            return JsonResponse({'ok': True, 'code': p.code, 'count': n})
        return JsonResponse({'ok': False, 'msg': 'Không đọc được số màu từ ảnh.'})
    if action == 'all_zero':
        updated, checked = 0, 0
        for p in Painting.objects.exclude(image='').filter(color_count=0):
            checked += 1
            n = _detect_color_count(os.path.join(settings.MEDIA_ROOT, p.image))
            if n:
                p.color_count = n
                p.save(update_fields=['color_count'])
                updated += 1
        return JsonResponse({'ok': True, 'updated': updated,
                             'msg': f'Đã cập nhật {updated}/{checked} mã tranh thiếu số màu.'})
    return JsonResponse({'ok': False, 'msg': 'Hành động không hợp lệ.'})


@csrf_exempt
@staff_required
def doc_so_mau(request):
    """AJAX: nhận 1 ảnh, trả số màu ước lượng (cho form khai báo mã tranh)."""
    if request.method != 'POST':
        return HttpResponseNotFound('POST only')
    up = request.FILES.get('image')
    if not up or not (up.content_type or '').startswith('image/'):
        return JsonResponse({'ok': False, 'count': 0})
    fss = FileSystemStorage()
    tmp = fss.save(f'_tmp_count_{datetime.now():%H-%M-%S}_{up.name}', up)
    try:
        n = _detect_color_count(os.path.join(settings.MEDIA_ROOT, tmp))
    finally:
        _remove_media(tmp)
    return JsonResponse({'ok': True, 'count': n})


@csrf_exempt
@login_required(login_url='/login')
def rot_yeu_cau_list(request):
    """JSON danh sách yêu cầu rót đang chờ. Nhân viên chỉ thấy việc giao cho mình
    hoặc cho 'mọi người'; quản lý thấy tất cả."""
    from pha.models import PourRequest
    qs = PourRequest.objects.filter(status=PourRequest.STATUS_PENDING)
    if not request.user.is_staff:
        from django.db.models import Q
        qs = qs.filter(Q(assignee='') | Q(assignee=request.user.username))
    pmap = _painting_map()
    rows = []
    for r in qs:
        t = r.created_time
        try:
            t = t.astimezone(_VN) if _VN else t
        except Exception:
            pass
        p = pmap.get(r.painting.strip().lower())
        rows.append({
            'id': r.id, 'painting': r.painting, 'size': r.size, 'qty': r.qty,
            'count': p.color_count if p else 0, 'note': r.note,
            'assignee': r.assignee, 'by': r.created_by,
            'image': ('/media/' + p.image) if (p and p.image) else '',
            'dt': t.strftime('%d/%m %H:%M'),
        })
    return JsonResponse({'rows': rows, 'count': len(rows)})


@csrf_exempt
@login_required(login_url='/login')
def lich_su_rot(request):
    """Lịch sử các lượt rót (mới nhất trước). ?range=today|all."""
    from pha.models import PourLog
    now = _now()
    if request.GET.get('range') == 'today':
        qs = PourLog.objects.filter(day=now.strftime('%Y-%m-%d'))
    else:
        qs = PourLog.objects.all()
    logs = list(qs.order_by('-created_time')[:100])
    # Cờ NGHI TRÙNG: cùng MÃ + cùng NGƯỜI, cách nhau <= 10 phút -> tô đỏ ở bảng để quản lý
    # dễ nhận ra & xoá ĐÚNG dòng thừa (tránh nhầm). Chỉ gợi ý nhìn, không tự xoá.
    def _is_dup(lg):
        for o in logs:
            if o.id != lg.id and o.painting == lg.painting and (o.user or '') == (lg.user or ''):
                try:
                    if abs((o.created_time - lg.created_time).total_seconds()) <= 600:
                        return True
                except Exception:
                    pass
        return False
    rows = []
    for log in logs:
        t = log.created_time
        try:
            t = t.astimezone(_VN) if _VN else t
        except Exception:
            pass
        rows.append({
            'id': log.id, 'dt': t.strftime('%d/%m %H:%M'), 'painting': log.painting,
            'size': log.size, 'qty': log.qty, 'colors': log.color_count, 'user': log.user or '',
            'dup': _is_dup(log),
        })
    return JsonResponse({'rows': rows})


@csrf_exempt
@staff_required
def xoa_lich_su_rot(request):
    """Xoá 1 dòng lịch sử rót, hoặc xoá TẤT CẢ (chỉ quản lý). Lưu ý: xoá khỏi cả thống kê."""
    if request.method != 'POST':
        return HttpResponseNotFound('POST only')
    from pha.models import PourLog
    if request.POST.get('all') == '1':
        PourLog.objects.all().delete()
        return JsonResponse({'ok': True, 'msg': 'Đã xoá toàn bộ lịch sử rót.'})
    pid = request.POST.get('id')
    if pid:
        n, _ = PourLog.objects.filter(id=pid).delete()
        return JsonResponse({'ok': bool(n), 'msg': 'Đã xoá 1 dòng.' if n else 'Không tìm thấy.'})
    return JsonResponse({'ok': False, 'msg': 'Thiếu tham số.'})


@csrf_exempt
@staff_required
def quan_ly_giao_rot(request):
    """AJAX cho APP QUẢN LÝ (điện thoại): giao việc rót / xoá / đánh dấu đã rót."""
    if request.method != 'POST':
        return HttpResponseNotFound('POST only')
    from pha.models import Painting, PourRequest
    action = request.POST.get('action')
    if action == 'add':
        code = (request.POST.get('painting') or '').strip()
        p = Painting.objects.filter(code__iexact=code).first()
        if not p:
            return JsonResponse({'ok': False, 'msg': f'Mã tranh "{code}" chưa có trong danh mục.'})
        try:
            qty = max(1, int(request.POST.get('qty') or 1))
        except ValueError:
            qty = 1
        req = PourRequest.objects.create(
            painting=p.code, size=_norm_size(request.POST.get('size')),
            colors=[], qty=qty,
            note=(request.POST.get('note') or '').strip(),
            assignee=(request.POST.get('assignee') or '').strip(),
            created_by=request.user.username,
        )
        from pha import push
        push.notify_pour(req)
        return JsonResponse({'ok': True, 'msg': f'Đã giao rót {p.code} ×{qty}'
                             + (f' ({req.size})' if req.size else '') + '.'})
    if action == 'delete':
        PourRequest.objects.filter(id=request.POST.get('id')).delete()
        return JsonResponse({'ok': True, 'msg': 'Đã xoá yêu cầu.'})
    if action == 'done':
        req = PourRequest.objects.filter(id=request.POST.get('id')).first()
        if req and req.status != PourRequest.STATUS_DONE:
            cc, _ = _painting_count(req.painting)
            _record_pour(req.painting, req.qty, cc, request.user.username, req, size=req.size)
        return JsonResponse({'ok': True, 'msg': 'Đã đánh dấu rót xong.'})
    return JsonResponse({'ok': False, 'msg': 'Hành động không hợp lệ.'})


@csrf_exempt
@login_required(login_url='/login')
def push_key(request):
    """Khoá công khai VAPID cho trình duyệt đăng ký Web Push ('' nếu chưa bật)."""
    from pha import push
    return JsonResponse({'key': push.public_key()})


@csrf_exempt
@login_required(login_url='/login')
def push_subscribe(request):
    """Lưu đăng ký Web Push của trình duyệt nhân viên (gắn với tài khoản đang đăng nhập)."""
    if request.method != 'POST':
        return HttpResponseNotFound('POST only')
    from pha.models import PushSubscription
    try:
        data = json.loads(request.body.decode('utf-8'))
    except (ValueError, TypeError, UnicodeDecodeError):
        return JsonResponse({'ok': False})
    endpoint = data.get('endpoint')
    keys = data.get('keys') or {}
    p256dh, auth = keys.get('p256dh'), keys.get('auth')
    if not endpoint or not p256dh or not auth:
        return JsonResponse({'ok': False})
    PushSubscription.objects.update_or_create(
        endpoint=endpoint,
        defaults={'username': request.user.username, 'p256dh': p256dh, 'auth': auth})
    return JsonResponse({'ok': True})


@csrf_exempt
@staff_required
def thong_ke_rot(request):
    label, agg = _pour_stats(request.GET.get('range', 'today'), request.GET.get('month'))
    return JsonResponse({'label': label, 'pours': agg['pours'],
                         'paintings': agg['paintings'],
                         'colors_total': agg['colors_total'],
                         'rows': agg['rows'], 'daily': agg['daily'],
                         'by_size': agg['by_size']})


@csrf_exempt
@staff_required
def export_thong_ke_rot_excel(request):
    """Xuất báo cáo rót màu ra Excel (.xlsx) theo khoảng thời gian."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    range_ = request.GET.get('range', 'month')
    month = request.GET.get('month')
    label, qs = _pour_stats_qs(range_, month)
    agg = _pour_aggregate(qs)

    wb = Workbook()
    head_fill = PatternFill('solid', fgColor='6F42C1')
    head_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')

    # Sheet 1: tổng theo mã tranh
    ws = wb.active
    ws.title = "Tong theo ma tranh"
    ws.append(["BÁO CÁO RÓT MÀU"])
    ws.append([label])
    ws.append([f"Số lượt rót: {agg['pours']}    Tổng số tranh: {agg['paintings']}"
               f"    Tổng số màu đã rót: {agg['colors_total']}"])
    ws.append(["Mã tranh", "Số lượt rót", "Tổng số tranh", "Số màu/tranh", "Tổng số màu"])
    for c in ws[4]:
        c.fill = head_fill; c.font = head_font; c.alignment = center
    for u in agg['rows']:
        ws.append([u['painting'], u['pours'], u['qty'], u['cc'], u['colors']])
    for col, w in zip('ABCDE', (20, 14, 14, 14, 14)):
        ws.column_dimensions[col].width = w

    # Sheet 2: SỐ LƯỢNG TRANH THEO KÍCH THƯỚC
    wss = wb.create_sheet("Theo kich thuoc")
    wss.append(["SỐ LƯỢNG TRANH ĐÃ SẢN XUẤT THEO KÍCH THƯỚC"])
    wss.append([label])
    wss.append(["Kích thước", "Số tranh", "Số lượt rót"])
    for c in wss[3]:
        c.fill = head_fill; c.font = head_font; c.alignment = center
    for u in agg['by_size']:
        wss.append([u['size'], u['paintings'], u['pours']])
    wss.append(["TỔNG", agg['paintings'], agg['pours']])
    for col, w in zip('ABC', (18, 14, 14)):
        wss.column_dimensions[col].width = w

    # Sheet 3: chi tiết từng lượt rót
    ws2 = wb.create_sheet("Chi tiet rot")
    ws2.append(["Ngày giờ", "Mã tranh", "Kích thước", "Số lượng", "Số màu", "Người rót"])
    for c in ws2[1]:
        c.fill = head_fill; c.font = head_font; c.alignment = center
    for log in qs.order_by('created_time'):
        t = log.created_time
        try:
            t = t.astimezone(_VN) if _VN else t
        except Exception:
            pass
        ws2.append([t.strftime('%d/%m/%Y %H:%M'), log.painting, log.size or '',
                    log.qty, log.color_count, log.user or ''])
    for col, w in zip('ABCDEF', (18, 18, 12, 10, 10, 16)):
        ws2.column_dimensions[col].width = w

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="bao_cao_rot_mau.xlsx"'
    wb.save(resp)
    return resp


# ===================== SẢN XUẤT TRANH (module quản lý tự điền) =====================
def _prod_day(raw):
    """Chuẩn hoá ngày 'YYYY-MM-DD' -> (day, month); rỗng/sai -> hôm nay."""
    raw = (raw or '').strip()
    try:
        d = datetime.strptime(raw, '%Y-%m-%d')
        return d.strftime('%Y-%m-%d'), d.strftime('%Y-%m')
    except ValueError:
        now = _now()
        return now.strftime('%Y-%m-%d'), now.strftime('%Y-%m')


def _prod_stats_qs(range_, month_param):
    """Trả (label, queryset PaintingProduction) theo khoảng today/week/month."""
    from pha.models import PaintingProduction
    now = _now()
    if range_ == 'week':
        d = now.date()
        monday = d - timedelta(days=d.weekday())
        sunday = monday + timedelta(days=6)
        qs = PaintingProduction.objects.filter(day__gte=monday.strftime('%Y-%m-%d'),
                                               day__lte=sunday.strftime('%Y-%m-%d'))
        label = f"Tuần này ({monday.strftime('%d/%m')} – {sunday.strftime('%d/%m/%Y')})"
    elif range_ == 'month':
        m = month_param or now.strftime('%Y-%m')
        qs = PaintingProduction.objects.filter(month=m)
        label = "Tháng " + _fmt_month(m)
    elif range_ == 'all':
        qs = PaintingProduction.objects.all()
        label = "Tất cả"
    else:
        qs = PaintingProduction.objects.filter(day=now.strftime('%Y-%m-%d'))
        label = "Hôm nay (" + now.strftime('%d/%m/%Y') + ")"
    return label, qs


def _prod_aggregate(qs):
    """Tổng hợp sản xuất: số lượt ghi, tổng số tranh; theo kích thước / mã tranh / ngày."""
    entries = 0
    total = 0
    sizeacc, dayacc = {}, {}
    for r in qs:
        entries += 1
        q = max(1, int(r.qty or 1))
        total += q
        sz = (r.size or '').strip() or '(chưa ghi)'
        sizeacc[sz] = sizeacc.get(sz, 0) + q
        dayacc[r.day] = dayacc.get(r.day, 0) + q
    by_size = sorted([{'size': k, 'qty': v} for k, v in sizeacc.items()], key=lambda x: -x['qty'])
    daily = []
    for k in sorted(dayacc.keys()):
        try:
            lbl = datetime.strptime(k, '%Y-%m-%d').strftime('%d/%m')
        except ValueError:
            lbl = k
        daily.append({'label': lbl, 'qty': dayacc[k]})
    return {'entries': entries, 'total': total, 'by_size': by_size, 'daily': daily}


@csrf_exempt
@staff_required
def san_xuat(request):
    """Module SẢN XUẤT TRANH: quản lý tự điền số lượng tranh thành phẩm + thống kê."""
    from pha.models import PaintingProduction

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            try:
                qty = max(1, int(request.POST.get('qty') or 1))
            except ValueError:
                qty = 1
            day, month = _prod_day(request.POST.get('day'))
            size = _norm_size(request.POST.get('size'))
            if not size:
                messages.error(request, 'Vui lòng nhập kích thước.')
            else:
                PaintingProduction.objects.create(
                    day=day, month=month, painting='', size=size, qty=qty,
                    note=(request.POST.get('note') or '').strip(),
                    user=request.user.username,
                )
                messages.info(request, f'Đã ghi nhận {qty} tranh ({size}).')
        elif action == 'delete':
            PaintingProduction.objects.filter(id=request.POST.get('id')).delete()
            messages.info(request, 'Đã xoá 1 dòng.')
        return redirect('/san-xuat')

    now = _now()
    entries = list(PaintingProduction.objects.all()[:80])
    for e in entries:
        try:
            t = e.created_time.astimezone(_VN) if _VN else e.created_time
        except Exception:
            t = e.created_time
        e.dt = t.strftime('%d/%m %H:%M')
    months = sorted(set(PaintingProduction.objects.values_list('month', flat=True)), reverse=True)
    stat_months = [{'value': m, 'label': _fmt_month(m)} for m in months]
    label, qs0 = _prod_stats_qs('today', None)
    agg = _prod_aggregate(qs0)
    return render(request, 'san_xuat.html', {
        'entries': entries,
        'paint_sizes': _paint_sizes(),
        'today': now.strftime('%Y-%m-%d'),
        'stat_months': stat_months, 'stat_label': label, 'stat_agg': agg,
    })


@csrf_exempt
@staff_required
def thong_ke_san_xuat(request):
    label, qs = _prod_stats_qs(request.GET.get('range', 'today'), request.GET.get('month'))
    agg = _prod_aggregate(qs)
    return JsonResponse({'label': label, 'entries': agg['entries'], 'total': agg['total'],
                         'by_size': agg['by_size'], 'daily': agg['daily']})


@csrf_exempt
@staff_required
def export_san_xuat_excel(request):
    """Xuất báo cáo SẢN XUẤT ra Excel theo khoảng thời gian."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    label, qs = _prod_stats_qs(request.GET.get('range', 'month'), request.GET.get('month'))
    agg = _prod_aggregate(qs)
    wb = Workbook()
    head_fill = PatternFill('solid', fgColor='0D6EFD')
    head_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')

    ws = wb.active
    ws.title = "Theo kich thuoc"
    ws.append(["SẢN XUẤT TRANH — THEO KÍCH THƯỚC"])
    ws.append([label])
    ws.append([f"Tổng số tranh: {agg['total']}    Số lượt ghi: {agg['entries']}"])
    ws.append(["Kích thước", "Số tranh"])
    for c in ws[4]:
        c.fill = head_fill; c.font = head_font; c.alignment = center
    for u in agg['by_size']:
        ws.append([u['size'], u['qty']])
    ws.append(["TỔNG", agg['total']])
    for col, w in zip('AB', (18, 14)):
        ws.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Chi tiet")
    ws3.append(["Ngày", "Kích thước", "Số lượng", "Người nhập", "Ghi chú"])
    for c in ws3[1]:
        c.fill = head_fill; c.font = head_font; c.alignment = center
    for r in qs.order_by('day', 'id'):
        ws3.append([r.day, r.size or '', r.qty, r.user or '', r.note or ''])
    for col, w in zip('ABCDE', (12, 14, 10, 14, 40)):
        ws3.column_dimensions[col].width = w

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="bao_cao_san_xuat.xlsx"'
    wb.save(resp)
    return resp


# ===================== NĂNG SUẤT & LƯƠNG KHOÁN NHÂN VIÊN =====================
def _period_filter(range_, month_param):
    """Trả (label, kwargs) lọc theo khoảng cho các model có trường day/month."""
    now = _now()
    if range_ == 'week':
        d = now.date()
        mon = d - timedelta(days=d.weekday())
        sun = mon + timedelta(days=6)
        return (f"Tuần này ({mon.strftime('%d/%m')} – {sun.strftime('%d/%m/%Y')})",
                {'day__gte': mon.strftime('%Y-%m-%d'), 'day__lte': sun.strftime('%Y-%m-%d')})
    if range_ == 'month':
        m = month_param or now.strftime('%Y-%m')
        return "Tháng " + _fmt_month(m), {'month': m}
    if range_ == 'all':
        return "Tất cả", {}
    return "Hôm nay (" + now.strftime('%d/%m/%Y') + ")", {'day': now.strftime('%Y-%m-%d')}


def _productivity(range_, month_param):
    """Sản lượng theo nhân viên trong khoảng (để quản lý lãi/lỗ). Trả (label, rows, totals).
    KHÔNG tính lương — lương tính theo NGÀY CÔNG ở phần mềm kế toán."""
    from pha.models import ProductionLog, PourLog, PaintingProduction
    label, f = _period_filter(range_, month_param)
    acc = {}

    def row(u):
        key = u or '(không tên)'
        return acc.setdefault(key, {'user': key, 'pha': 0, 'rot_p': 0, 'rot_c': 0, 'sx': 0})

    for log in ProductionLog.objects.filter(**f):
        row(log.user)['pha'] += 1
    for log in PourLog.objects.filter(**f):
        q = max(1, int(log.qty or 1))
        r = row(log.user)
        r['rot_p'] += q
        r['rot_c'] += int(log.color_count or 0) * q
    for p in PaintingProduction.objects.filter(**f):
        row(p.user)['sx'] += max(1, int(p.qty or 1))

    rows = []
    for v in acc.values():
        v['out'] = v['pha'] + v['rot_p'] + v['sx']  # tổng đầu việc (xếp hạng)
        rows.append(v)
    rows.sort(key=lambda x: -x['out'])
    totals = {k: sum(r[k] for r in rows) for k in ('pha', 'rot_p', 'rot_c', 'sx', 'out')}
    return label, rows, totals


@csrf_exempt
@staff_required
def nang_suat(request):
    """Báo cáo SẢN LƯỢNG theo nhân viên (để quản lý lãi/lỗ). Lương tính ở phần kế toán."""
    from pha.models import ProductionLog, PourLog, PaintingProduction
    label, rows, totals = _productivity('month', None)
    months = set(ProductionLog.objects.values_list('month', flat=True))
    months |= set(PourLog.objects.values_list('month', flat=True))
    months |= set(PaintingProduction.objects.values_list('month', flat=True))
    stat_months = [{'value': m, 'label': _fmt_month(m)} for m in sorted(months, reverse=True)]
    return render(request, 'nang_suat.html', {
        'label': label, 'rows': rows, 'totals': totals,
        'stat_months': stat_months,
    })


@csrf_exempt
@staff_required
def thong_ke_nang_suat(request):
    label, rows, totals = _productivity(request.GET.get('range', 'month'),
                                        request.GET.get('month'))
    return JsonResponse({'label': label, 'rows': rows, 'totals': totals})


@csrf_exempt
@staff_required
def export_nang_suat_excel(request):
    """Xuất bảng SẢN LƯỢNG theo nhân viên ra Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    label, rows, totals = _productivity(request.GET.get('range', 'month'),
                                        request.GET.get('month'))
    wb = Workbook()
    ws = wb.active
    ws.title = "San luong"
    head_fill = PatternFill('solid', fgColor='6610F2')
    head_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')
    ws.append(["BẢNG SẢN LƯỢNG THEO NHÂN VIÊN"])
    ws.append([label])
    ws.append(["Nhân viên", "Mẻ pha", "Tranh rót", "Màu rót", "Tranh SX", "Tổng đầu việc"])
    for c in ws[3]:
        c.fill = head_fill; c.font = head_font; c.alignment = center
    for r in rows:
        ws.append([r['user'], r['pha'], r['rot_p'], r['rot_c'], r['sx'], r['out']])
    ws.append(["TỔNG", totals['pha'], totals['rot_p'], totals['rot_c'], totals['sx'], totals['out']])
    for col, w in zip('ABCDEF', (18, 10, 12, 12, 12, 16)):
        ws.column_dimensions[col].width = w
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="san_luong.xlsx"'
    wb.save(resp)
    return resp


# ===================== DOANH THU – LỢI NHUẬN =====================
def _price_of(size):
    from pha.models import AppSetting
    try:
        return float((AppSetting.get('PRICE_' + (size or ''), '0') or '0').replace(',', '').strip() or 0)
    except (TypeError, ValueError):
        return 0.0


def _price_list():
    """Danh sách [{size, price}] cho form giá bán: khổ gợi ý + khổ đã có trong dữ liệu SX."""
    from pha.models import PaintingProduction
    sizes = list(_paint_sizes())
    for s in sorted(set(PaintingProduction.objects.exclude(size='').values_list('size', flat=True))):
        if s not in sizes:
            sizes.append(s)
    return [{'size': s, 'price': round(_price_of(s))} for s in sizes]


def _labor_per_day():
    """Đơn giá 1 ngày công (đồng/ngày) lưu trong AppSetting — để ước tính nhân công cho lãi/lỗ."""
    from pha.models import AppSetting
    try:
        return float((AppSetting.get('LABOR_PER_DAY', '0') or '0').replace(',', '').strip() or 0)
    except (TypeError, ValueError):
        return 0.0


def _profit(range_, month_param):
    """Doanh thu (SX theo khổ × giá bán) − chi phí sơn − nhân công (ngày công × đơn giá) = lợi nhuận."""
    from pha.models import PaintingProduction, ProductionLog, Attendance
    label, f = _period_filter(range_, month_param)
    sizeqty = {}
    for p in PaintingProduction.objects.filter(**f):
        sz = (p.size or '').strip() or '(chưa ghi)'
        sizeqty[sz] = sizeqty.get(sz, 0) + max(1, int(p.qty or 1))
    rows, revenue = [], 0.0
    for sz, q in sizeqty.items():
        price = _price_of(sz)
        rev = q * price
        revenue += rev
        rows.append({'size': sz, 'qty': q, 'price': round(price), 'revenue': round(rev)})
    rows.sort(key=lambda x: -x['revenue'])
    paint_cost = ProductionLog.objects.filter(**f).aggregate(s=Sum('cost'))['s'] or 0
    # Nhân công ước tính theo NGÀY CÔNG (khớp cách trả lương theo ngày công ở kế toán)
    work_days = Attendance.objects.filter(check_in__isnull=False, **f).count()
    per_day = _labor_per_day()
    labor = work_days * per_day
    summary = {'revenue': round(revenue), 'paint_cost': round(paint_cost),
               'work_days': work_days, 'per_day': round(per_day),
               'labor': round(labor), 'profit': round(revenue - paint_cost - labor)}
    return label, rows, summary


@csrf_exempt
@staff_required
def loi_nhuan(request):
    """Báo cáo DOANH THU – LỢI NHUẬN (chỉ quản lý). Cấu hình giá bán theo kích thước."""
    from pha.models import AppSetting, ProductionLog, PaintingProduction
    if request.method == 'POST' and request.POST.get('action') == 'save_prices':
        for sz, pr in zip(request.POST.getlist('psize'), request.POST.getlist('pprice')):
            sz = (sz or '').strip()
            if sz:
                AppSetting.set('PRICE_' + sz, (pr or '0').replace(',', '').strip() or '0')
        AppSetting.set('LABOR_PER_DAY', (request.POST.get('labor_per_day') or '0').replace(',', '').strip() or '0')
        messages.info(request, 'Đã lưu giá bán & đơn giá ngày công.')
        return redirect('/loi-nhuan')

    label, rows, summary = _profit('month', None)
    months = set(ProductionLog.objects.values_list('month', flat=True))
    months |= set(PaintingProduction.objects.values_list('month', flat=True))
    stat_months = [{'value': m, 'label': _fmt_month(m)} for m in sorted(months, reverse=True)]
    return render(request, 'loi_nhuan.html', {
        'label': label, 'rows': rows, 'summary': summary,
        'price_list': _price_list(), 'stat_months': stat_months,
        'labor_per_day': round(_labor_per_day()),
    })


@csrf_exempt
@staff_required
def thong_ke_loi_nhuan(request):
    label, rows, summary = _profit(request.GET.get('range', 'month'), request.GET.get('month'))
    return JsonResponse({'label': label, 'rows': rows, 'summary': summary})


@csrf_exempt
@staff_required
def export_loi_nhuan_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    label, rows, summary = _profit(request.GET.get('range', 'month'), request.GET.get('month'))
    wb = Workbook()
    ws = wb.active
    ws.title = "Loi nhuan"
    head_fill = PatternFill('solid', fgColor='0D6EFD')
    head_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')
    ws.append(["BÁO CÁO DOANH THU – LỢI NHUẬN"])
    ws.append([label])
    ws.append(["Doanh thu", summary['revenue']])
    ws.append(["Chi phí sơn", summary['paint_cost']])
    ws.append([f"Nhân công ({summary.get('work_days', 0)} công × {summary.get('per_day', 0):,}đ)",
               summary['labor']])
    ws.append(["LỢI NHUẬN", summary['profit']])
    ws.append([])
    ws.append(["Kích thước", "Số tranh", "Giá bán (đ)", "Doanh thu (đ)"])
    for c in ws[8]:
        c.fill = head_fill; c.font = head_font; c.alignment = center
    for r in rows:
        ws.append([r['size'], r['qty'], r['price'], r['revenue']])
    for col, w in zip('ABCD', (18, 12, 16, 18)):
        ws.column_dimensions[col].width = w
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="loi_nhuan.xlsx"'
    wb.save(resp)
    return resp


# ===================== CHẤM CÔNG (theo IP Wifi công ty) =====================
def _client_ip(request):
    """IP thật của máy nhân viên. Sau nginx, lấy entry CUỐI của X-Forwarded-For
    (do nginx tự thêm = IP thật, chống giả mạo); không có thì REMOTE_ADDR."""
    xff = request.META.get('HTTP_X_FORWARDED_FOR', '')
    if xff:
        parts = [p.strip() for p in xff.split(',') if p.strip()]
        if parts:
            return parts[-1]
    return request.META.get('REMOTE_ADDR', '') or ''


def _attendance_ips():
    from pha.models import AppSetting
    raw = (AppSetting.get('ATTENDANCE_IPS', '') or '')
    return [s.strip() for s in raw.replace('\n', ',').split(',') if s.strip()]


def _ip_allowed(ip):
    """True nếu ip khớp danh sách IP công ty (nhập tay) HOẶC IP xưởng TỰ CẬP NHẬT
    (pha/wifi_ip.py — agent in ở xưởng báo về, kèm ân hạn IP cũ). Mỗi mục nhập tay:
    IP đầy đủ (113.161.4.20) hoặc tiền tố theo octet (113.161. / 113.161.4).
    Chưa cấu hình GÌ CẢ -> cho phép (để lấy IP)."""
    allow = _attendance_ips()
    try:
        from pha.wifi_ip import auto_ips
        auto = auto_ips()
    except Exception:
        auto = []
    if not allow and not auto:
        return True
    if ip and ip in auto:
        return True
    for a in allow:
        if ip == a:
            return True
        if a.endswith('.') and ip.startswith(a):
            return True
        if ip.startswith(a + '.'):
            return True
    return False


def _ip_private(ip):
    """True nếu ip rỗng/nội bộ/localhost -> dấu hiệu nginx CHƯA chuyển IP thật."""
    if not ip:
        return True
    if ip.startswith(('127.', '10.', '192.168.', '169.254.', '::1')):
        return True
    if ip.startswith('172.'):
        try:
            return 16 <= int(ip.split('.')[1]) <= 31
        except (IndexError, ValueError):
            return False
    return False


@csrf_exempt
@login_required(login_url='/login')
def cham_cong_ip(request):
    """Trang kiểm tra: server nhận IP nào cho máy bạn (mở trên điện thoại để kiểm)."""
    ip = _client_ip(request)
    company = _attendance_ips()
    data = {
        'ip': ip, 'allowed': _ip_allowed(ip), 'configured': bool(company),
        'private': _ip_private(ip), 'company_ips': company,
        'x_forwarded_for': request.META.get('HTTP_X_FORWARDED_FOR', ''),
        'remote_addr': request.META.get('REMOTE_ADDR', ''),
    }
    if request.GET.get('json') == '1':
        return JsonResponse(data)
    return render(request, 'cham_cong_ip.html', {**data, 'is_staff': request.user.is_staff})


def _vn_dt(dt):
    if not dt:
        return None
    try:
        return dt.astimezone(_VN) if _VN else dt
    except Exception:
        return dt


def _hm(dt):
    d = _vn_dt(dt)
    return d.strftime('%H:%M') if d else ''


def _fmt_day(d):
    try:
        return datetime.strptime(d, '%Y-%m-%d').strftime('%d/%m')
    except ValueError:
        return d


def _mk_dt(day_str, hm_str):
    """Ghép ngày 'YYYY-MM-DD' + giờ 'HH:MM' -> datetime có timezone VN. None nếu giờ trống/sai."""
    hm = (hm_str or '').strip()
    if not hm:
        return None
    try:
        d = datetime.strptime((day_str or '').strip() + ' ' + hm, '%Y-%m-%d %H:%M')
    except ValueError:
        return None
    return d.replace(tzinfo=_VN) if _VN else d


def _att_hours(rec):
    if rec.check_in and rec.check_out:
        h = (rec.check_out - rec.check_in).total_seconds() / 3600.0
        return round(h, 2) if h > 0 else 0
    return 0


_WEEKDAYS = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']   # 0=T2 .. 6=CN (Python weekday)


def _hm_to_min(hhmm):
    try:
        h, m = (hhmm or '').split(':')
        return int(h) * 60 + int(m)
    except (ValueError, AttributeError):
        return None


def _att_cfg():
    """Cấu hình khung giờ & quy định chấm công (AppSetting)."""
    from pha.models import AppSetting

    def num(k, d=0):
        try:
            return float((AppSetting.get(k, str(d)) or str(d)).replace(',', '').strip() or d)
        except (TypeError, ValueError):
            return float(d)
    days_raw = AppSetting.get('WORK_DAYS', '0,1,2,3,4,5') or '0,1,2,3,4,5'
    workdays = set(int(x) for x in days_raw.split(',') if x.strip().isdigit())
    return {
        'start': AppSetting.get('WORK_START', '08:00') or '08:00',
        'end': AppSetting.get('WORK_END', '17:00') or '17:00',
        'workdays': workdays or {0, 1, 2, 3, 4, 5},
        'grace': int(num('LATE_GRACE', 0)),
        'fine_min': num('LATE_FINE_PER_MIN', 0),
        'fine_fixed': num('LATE_FINE_FIXED', 0),
        'ot_rate': num('OT_RATE', 0),
        'ot_min': num('OT_MIN_HOURS', 1),   # tăng ca phải đạt ngưỡng này (giờ) mới được tính
        'hol_off': {d.strip() for d in (AppSetting.get('HOLIDAYS_OFF', '') or '').split(',') if d.strip()},
        'hol_extra': _parse_holidays_extra(AppSetting.get('HOLIDAYS_EXTRA', '')),
        # Nghỉ trưa + chuông báo dậy (chuông kêu TRƯỚC giờ hết nghỉ trưa N phút)
        'lunch_start': AppSetting.get('LUNCH_START', '12:00') or '12:00',
        'lunch_end': AppSetting.get('LUNCH_END', '13:30') or '13:30',
        'bell_before': int(num('BELL_BEFORE_MIN', 3)),
        # Chuông BẮT BUỘC (mặc định): mọi máy phải kêu, không tắt được; 0 = cho từng máy tự chọn
        'bell_mandatory': (AppSetting.get('BELL_MANDATORY', '1') or '1') == '1',
    }


def _parse_holidays_extra(raw):
    """'YYYY-MM-DD=Tên, YYYY-MM-DD' -> {date: name}. Ngày tự thêm của công ty."""
    out = {}
    for item in (raw or '').split(','):
        item = item.strip()
        if not item:
            continue
        if '=' in item:
            dd, nm = item.split('=', 1)
            out[dd.strip()] = nm.strip() or 'Nghỉ lễ'
        else:
            out[item] = 'Nghỉ lễ'
    return out


def _holiday_lookup(day, cfg):
    """Tên ngày lễ của 'YYYY-MM-DD' theo cfg (đã gồm ngày tự thêm/bỏ). '' nếu ngày thường."""
    if day in cfg.get('hol_off', ()):
        return ''
    ex = cfg.get('hol_extra', {})
    if day in ex:
        return ex[day]
    import pha.holidays_vn as _hol
    return _hol.name_of(day)


def _holiday_name(day):
    """Tên ngày lễ VN của 'YYYY-MM-DD' (dùng ngoài _att_calc). '' nếu ngày thường."""
    return _holiday_lookup(day, _att_cfg())


def _nudge_key_for_view():
    try:
        import pha.attend_nudge as _n
        return _n._nudge_key()
    except Exception:
        return ''


def _upcoming_holidays(now, cfg, limit=24):
    """Danh sách ngày nghỉ lễ sắp tới (năm nay + năm sau), đã áp ngày tự thêm/bỏ."""
    import pha.holidays_vn as _hol
    today = now.strftime('%Y-%m-%d')
    merged = {}
    merged.update(_hol.holidays(now.year))
    merged.update(_hol.holidays(now.year + 1))
    merged.update(cfg.get('hol_extra', {}))
    for d in cfg.get('hol_off', ()):
        merged.pop(d, None)
    extra = cfg.get('hol_extra', {})
    out = []
    for k, v in sorted(merged.items()):
        if k < today:
            continue
        try:
            df = datetime.strptime(k, '%Y-%m-%d').strftime('%d/%m/%Y')
        except ValueError:
            df = k
        out.append({'date': k, 'date_fmt': df, 'name': v, 'custom': k in extra})
    return out[:limit]


def _att_calc(rec, cfg):
    """Tính cho 1 ngày công: giờ làm, phút đi muộn, giờ tăng ca, tiền phạt, tiền OT."""
    out = {'hours': _att_hours(rec), 'late_min': 0, 'ot_hours': 0.0,
           'fine': 0, 'ot_pay': 0, 'workday': True, 'weekday': '', 'holiday': ''}
    if not rec.check_in:
        out['holiday'] = _holiday_lookup(rec.day, cfg)
        return out
    cin = _vn_dt(rec.check_in)
    wd = cin.weekday()
    out['weekday'] = _WEEKDAYS[wd]
    out['holiday'] = _holiday_lookup(rec.day, cfg)
    out['workday'] = (wd in cfg['workdays']) and not out['holiday']
    start, end = _hm_to_min(cfg['start']), _hm_to_min(cfg['end'])
    cin_min = cin.hour * 60 + cin.minute
    cout = _vn_dt(rec.check_out)
    cout_min = (cout.hour * 60 + cout.minute) if cout else None
    if out['workday']:
        if start is not None and cin_min > start + cfg['grace']:
            out['late_min'] = cin_min - (start + cfg['grace'])
            out['fine'] = round(out['late_min'] * cfg['fine_min']
                                + (cfg['fine_fixed'] if out['late_min'] > 0 else 0))
        if cout_min is not None and end is not None and cout_min > end:
            out['ot_hours'] = round((cout_min - end) / 60.0, 2)
    else:
        out['ot_hours'] = out['hours']     # làm ngày nghỉ -> tính tăng ca toàn bộ
    # Chỉ tính tăng ca khi đạt ngưỡng tối thiểu (vd 1 giờ) — lố vài phút không tính
    if out['ot_hours'] < cfg.get('ot_min', 0):
        out['ot_hours'] = 0.0
    out['ot_pay'] = round(out['ot_hours'] * cfg['ot_rate'])
    return out


def _emp_cong(user, **f):
    """Ngày công 1 nhân viên trong khoảng f (day=.. hoặc month=..):
    số ngày công, tổng giờ làm, giờ tăng ca, số phút đi muộn."""
    from pha.models import Attendance
    cfg = _att_cfg()
    days = late_min = 0
    hours = ot = 0.0
    for r in Attendance.objects.filter(user=user, **f):
        c = _att_calc(r, cfg)
        if r.check_in:
            days += 1
        hours += c['hours']; ot += c['ot_hours']; late_min += c['late_min']
    return {'days': days, 'hours': round(hours, 1), 'ot': round(ot, 1), 'late_min': late_min}


@csrf_exempt
@login_required(login_url='/login')
def cham_cong(request):
    """Trang CHẤM CÔNG cho nhân viên: bấm Vào làm / Tan làm (chỉ khi đúng Wifi công ty)."""
    from pha.models import Attendance
    now = _now()
    ip = _client_ip(request)
    ip_ok = _ip_allowed(ip)
    configured = bool(_attendance_ips())

    if request.method == 'POST':
        action = request.POST.get('action')
        if not ip_ok:
            return JsonResponse({'ok': False, 'wifi': False,
                                 'msg': 'Bạn cần kết nối Wifi công ty để chấm công.', 'ip': ip})
        # ----- KHOÁ THIẾT BỊ 1-1 (chống chấm hộ) -----
        # Mã thiết bị ưu tiên lấy từ COOKIE server cấp (bền — không bị iOS xoá sau 7 ngày
        # như localStorage). localStorage chỉ dùng để DI CƯ token cũ -> tránh khoá nhầm
        # hàng loạt sau khi deploy.
        from pha.models import DeviceBind
        import secrets
        u = request.user.username
        cookie_dev = (request.COOKIES.get('dali_dev') or '').strip()
        ls_dev = (request.POST.get('device') or '').strip()   # localStorage (bản cũ)
        mine = DeviceBind.objects.filter(username=u).first()

        set_cookie_to = None
        if mine and cookie_dev and cookie_dev == mine.token:
            device = cookie_dev                       # đã có cookie bền khớp khoá
        elif mine and ls_dev and ls_dev == mine.token:
            device = ls_dev                           # token cũ trong localStorage -> di cư vào cookie
            set_cookie_to = ls_dev
        elif mine:
            device = cookie_dev or ls_dev or ''       # không khớp -> sẽ bị chặn bên dưới (khác máy thật)
        else:
            device = cookie_dev or ls_dev or ('d-' + secrets.token_hex(16))   # máy mới
            set_cookie_to = device
        if cookie_dev != device and device:
            set_cookie_to = device                    # đồng bộ cookie = token đang dùng

        def _dev_resp(payload, status=200):
            r = JsonResponse(payload, status=status)
            if set_cookie_to:
                r.set_cookie('dali_dev', set_cookie_to, max_age=63072000,   # ~2 năm
                             samesite='Lax', secure=request.is_secure())
            return r

        if not device:
            return _dev_resp({'ok': False, 'msg': 'Không lấy được mã thiết bị. Tắt chế độ ẩn danh / bật JavaScript rồi thử lại.'})
        other = DeviceBind.objects.filter(token=device).exclude(username=u).first()
        if other:
            return _dev_resp({'ok': False, 'msg': f'Điện thoại này đã gắn với tài khoản "{other.username}". Mỗi máy chỉ chấm công cho 1 người.'})
        if mine and mine.token != device:
            return _dev_resp({'ok': False, 'msg': 'Không nhận diện được thiết bị này (có thể trình duyệt vừa xoá dữ liệu). Báo quản lý mở khoá thiết bị giúp bạn (1 lần).'})
        if not mine:
            DeviceBind.objects.create(username=u, token=device,
                                      user_agent=(request.META.get('HTTP_USER_AGENT', '') or '')[:200])
        rec, _ = Attendance.objects.get_or_create(
            user=u, day=now.strftime('%Y-%m-%d'), defaults={'month': now.strftime('%Y-%m')})
        if action == 'in':
            if rec.check_in:
                return _dev_resp({'ok': False, 'msg': 'Bạn đã chấm công VÀO hôm nay rồi.',
                                  'in': _hm(rec.check_in), 'out': _hm(rec.check_out)})
            rec.check_in = now; rec.ip_in = ip; rec.device_in = device; rec.save()
            import pha.motivate as _mot
            return _dev_resp({'ok': True, 'msg': 'Đã chấm công VÀO lúc ' + _hm(now),
                              'in': _hm(rec.check_in), 'out': _hm(rec.check_out),
                              'motivate': _mot.quote_for(now)})
        if action == 'undo_out':
            # Hoàn tác TAN LÀM bấm nhầm — chỉ trong 15 phút sau khi bấm
            if not rec.check_out:
                return _dev_resp({'ok': False, 'msg': 'Hôm nay bạn chưa bấm tan làm.'})
            if (now - rec.check_out).total_seconds() > 15 * 60:
                return _dev_resp({'ok': False, 'msg': 'Quá 15 phút — nhờ quản lý sửa giờ ra giúp.'})
            rec.check_out = None; rec.ip_out = ''; rec.device_out = ''; rec.save()
            return _dev_resp({'ok': True, 'msg': 'Đã hoàn tác tan làm — tiếp tục làm việc nhé! 💪',
                              'in': _hm(rec.check_in), 'out': '', 'undone': True})
        if action == 'out':
            if not rec.check_in:
                return _dev_resp({'ok': False, 'msg': 'Bạn chưa chấm công VÀO.'})
            # CHỐNG BẤM NHẦM: tan làm sớm bất thường / tan làm lần 2 -> phải xác nhận rõ
            if request.POST.get('confirm') != '1':
                if rec.check_out:
                    return _dev_resp({'ok': False, 'need_confirm': 'update',
                                      'msg': f'Bạn đã tan làm lúc {_hm(rec.check_out)}. '
                                             f'Cập nhật giờ ra thành {_hm(now)}?'})
                mins = int((now - rec.check_in).total_seconds() // 60)
                if mins < 60:
                    return _dev_resp({'ok': False, 'need_confirm': 'early', 'mins': mins,
                                      'msg': f'Bạn mới VÀO LÀM {mins} phút trước. '
                                             f'Chắc chắn TAN LÀM bây giờ?'})
            rec.check_out = now; rec.ip_out = ip; rec.device_out = device; rec.save()
            day_s, month_s = now.strftime('%Y-%m-%d'), now.strftime('%Y-%m')
            tc = _att_calc(rec, _att_cfg())  # công hôm nay
            mc = _emp_cong(u, month=month_s)  # tổng tháng
            payload = {'ok': True, 'msg': 'Đã chấm công RA lúc ' + _hm(now),
                       'in': _hm(rec.check_in), 'out': _hm(rec.check_out),
                       'cong': {'time': _hm(now),
                                'today_hours': round(tc['hours'], 1),
                                'today_ot': round(tc['ot_hours'], 1),
                                'today_late': tc['late_min'],
                                'month_days': mc['days'], 'month_hours': mc['hours'],
                                'month_ot': mc['ot'], 'month_late': mc['late_min'],
                                'month_label': _fmt_month(month_s)}}
            # Lương CHÍNH XÁC kéo từ phần mềm kế toán (nếu đã cấu hình + gọi được)
            sal = _ketoan_salary(u, day_s, month_s)
            if sal and 'error' not in sal:
                payload['salary'] = sal
            return _dev_resp(payload)
        return _dev_resp({'ok': False, 'msg': 'Hành động không hợp lệ.'})

    today = Attendance.objects.filter(user=request.user.username,
                                      day=now.strftime('%Y-%m-%d')).first()
    recent = [{'day': _fmt_day(r.day), 'in': _hm(r.check_in), 'out': _hm(r.check_out),
               'hours': _att_hours(r)}
              for r in Attendance.objects.filter(user=request.user.username).order_by('-day')[:14]]
    from pha.models import LeaveRequest
    _lv_label = {'pending': 'Chờ duyệt', 'approved': 'Đã duyệt', 'rejected': 'Từ chối'}
    my_leaves = [{'rng': _fmt_day(lr.from_day) + ('' if lr.from_day == lr.to_day
                                                  else ' → ' + _fmt_day(lr.to_day)),
                  'reason': lr.reason, 'status': lr.status,
                  'status_label': _lv_label.get(lr.status, lr.status)}
                 for lr in LeaveRequest.objects.filter(user=request.user.username)[:6]]
    # Trạng thái hôm nay + tổng tháng (cho thẻ trạng thái & mini-thống kê)
    if today and today.check_out:
        state = 'done'
    elif today and today.check_in:
        state = 'working'
    else:
        state = 'not_in'
    cfg = _att_cfg()
    msum = _emp_cong(request.user.username, month=now.strftime('%Y-%m'))
    return render(request, 'cham_cong.html', {
        'today_in': _hm(today.check_in) if today else '',
        'today_out': _hm(today.check_out) if today else '',
        'ip': ip, 'ip_ok': ip_ok, 'configured': configured, 'recent': recent,
        'today_label': now.strftime('%d/%m/%Y'),
        'today_iso': now.strftime('%Y-%m-%d'),
        'my_leaves': my_leaves,
        'state': state,
        'month_label': _fmt_month(now.strftime('%Y-%m')),
        'm_days': msum['days'], 'm_hours': msum['hours'],
        'm_ot': msum['ot'], 'm_late': msum['late_min'],
        'lunch_start': cfg['lunch_start'], 'lunch_end': cfg['lunch_end'],
        'work_start': cfg['start'], 'work_end': cfg['end'],
    })


@csrf_exempt
@staff_required
def cham_cong_quan_ly(request):
    """Quản lý CHẤM CÔNG: IP Wifi + bảng công + KHOÁ THIẾT BỊ + CẢNH BÁO bất thường."""
    from pha.models import Attendance, AppSetting, DeviceBind
    if request.method == 'POST':
        act = request.POST.get('action')
        if act == 'save_ips':
            AppSetting.set('ATTENDANCE_IPS', (request.POST.get('ips') or '').strip())
            messages.info(request, 'Đã lưu IP Wifi công ty.')
            return redirect('/cham-cong-quan-ly')
        if act == 'save_schedule':
            AppSetting.set('WORK_START', (request.POST.get('work_start') or '08:00').strip())
            AppSetting.set('WORK_END', (request.POST.get('work_end') or '17:00').strip())
            AppSetting.set('LUNCH_START', (request.POST.get('lunch_start') or '12:00').strip())
            AppSetting.set('LUNCH_END', (request.POST.get('lunch_end') or '13:30').strip())
            days = [d for d in request.POST.getlist('workday') if d.isdigit()]
            AppSetting.set('WORK_DAYS', ','.join(days) if days else '0,1,2,3,4,5')
            for key, fld in (('LATE_GRACE', 'grace'), ('LATE_FINE_PER_MIN', 'fine_min'),
                             ('LATE_FINE_FIXED', 'fine_fixed'), ('OT_RATE', 'ot_rate'),
                             ('OT_MIN_HOURS', 'ot_min')):
                AppSetting.set(key, (request.POST.get(fld) or '0').replace(',', '').strip() or '0')
            AppSetting.set('BELL_BEFORE_MIN',
                           (request.POST.get('bell_before') or '3').replace(',', '').strip() or '3')
            # select luôn được gửi kèm form -> form cũ thiếu field thì giữ mặc định 'bắt buộc'
            AppSetting.set('BELL_MANDATORY',
                           '0' if request.POST.get('bell_mode') == 'optional' else '1')
            messages.info(request, 'Đã lưu khung giờ & quy định.')
            return redirect('/cham-cong-quan-ly')
        if act == 'reset_device':
            u = (request.POST.get('user') or '').strip()
            DeviceBind.objects.filter(username=u).delete()
            messages.info(request, f'Đã reset thiết bị cho "{u}". Lần chấm tiếp theo của họ sẽ gắn với máy mới.')
            return redirect('/cham-cong-quan-ly')
        if act == 'save_api_key':
            AppSetting.set('KETOAN_API_KEY', (request.POST.get('api_key') or '').strip())
            messages.info(request, 'Đã lưu khoá API kế toán/lương.')
            return redirect('/cham-cong-quan-ly')
        if act == 'save_ketoan_pull':
            AppSetting.set('KETOAN_SALARY_URL', (request.POST.get('ketoan_url') or '').strip())
            AppSetting.set('KETOAN_PULL_KEY', (request.POST.get('ketoan_pull_key') or '').strip())
            messages.info(request, 'Đã lưu kết nối kéo lương từ kế toán.')
            return redirect('/cham-cong-quan-ly')
        if act == 'save_prod_feed':
            AppSetting.set('KETOAN_PROD_URL', (request.POST.get('prod_url') or '').strip())
            AppSetting.set('KETOAN_PROD_KEY', (request.POST.get('prod_key') or '').strip())
            messages.info(request, 'Đã lưu kết nối đẩy năng suất sang kế toán.')
            return redirect('/cham-cong-quan-ly')
        if act == 'save_attendance':
            au = (request.POST.get('att_user') or '').strip()
            ad = (request.POST.get('att_day') or '').strip()
            try:
                ddmm = datetime.strptime(ad, '%Y-%m-%d')
            except ValueError:
                ddmm = None
            if not au or not ddmm:
                messages.error(request, 'Cần chọn nhân viên và ngày hợp lệ.')
                return redirect('/cham-cong-quan-ly')
            cin = _mk_dt(ad, request.POST.get('att_in'))
            cout = _mk_dt(ad, request.POST.get('att_out'))
            if cin and cout and cout <= cin:
                messages.error(request, 'Giờ ra phải sau giờ vào.')
                return redirect('/cham-cong-quan-ly?month=' + ad[:7])
            rec, _c = Attendance.objects.get_or_create(user=au, day=ad, defaults={'month': ad[:7]})
            rec.month = ad[:7]
            rec.check_in = cin
            rec.check_out = cout
            rec.save()
            messages.info(request, f'Đã lưu công {au} ngày {ddmm.strftime("%d/%m/%Y")} (sửa thủ công).')
            return redirect('/cham-cong-quan-ly?month=' + ad[:7])
        if act == 'delete_attendance':
            au = (request.POST.get('att_user') or '').strip()
            ad = (request.POST.get('att_day') or '').strip()
            Attendance.objects.filter(user=au, day=ad).delete()
            messages.info(request, f'Đã xoá chấm công {au} ngày {_fmt_day(ad)}.')
            return redirect('/cham-cong-quan-ly?month=' + (ad[:7] if len(ad) >= 7 else ''))
        if act == 'add_holiday':
            d = (request.POST.get('hd_date') or '').strip()
            nm = (request.POST.get('hd_name') or '').strip() or 'Nghỉ lễ'
            try:
                datetime.strptime(d, '%Y-%m-%d'); okd = True
            except ValueError:
                okd = False
            if not okd:
                messages.error(request, 'Ngày nghỉ không hợp lệ.')
                return redirect('/cham-cong-quan-ly')
            ex = _parse_holidays_extra(AppSetting.get('HOLIDAYS_EXTRA', ''))
            ex[d] = nm
            AppSetting.set('HOLIDAYS_EXTRA', ', '.join('%s=%s' % (k, v) for k, v in sorted(ex.items())))
            off = {x.strip() for x in (AppSetting.get('HOLIDAYS_OFF', '') or '').split(',') if x.strip()}
            off.discard(d)
            AppSetting.set('HOLIDAYS_OFF', ', '.join(sorted(off)))
            messages.info(request, f'Đã thêm ngày nghỉ {_fmt_day(d)} ({nm}).')
            return redirect('/cham-cong-quan-ly')
        if act == 'del_holiday':
            d = (request.POST.get('hd_date') or '').strip()
            ex = _parse_holidays_extra(AppSetting.get('HOLIDAYS_EXTRA', ''))
            if d in ex:
                ex.pop(d, None)
                AppSetting.set('HOLIDAYS_EXTRA', ', '.join('%s=%s' % (k, v) for k, v in sorted(ex.items())))
            else:
                off = {x.strip() for x in (AppSetting.get('HOLIDAYS_OFF', '') or '').split(',') if x.strip()}
                off.add(d)
                AppSetting.set('HOLIDAYS_OFF', ', '.join(sorted(off)))
            messages.info(request, f'Đã bỏ ngày nghỉ {_fmt_day(d)} (coi là ngày làm).')
            return redirect('/cham-cong-quan-ly')
        if act == 'leave_decide':
            from pha.models import LeaveRequest
            from pha.extra_views import _notify_leave_decision
            lr = LeaveRequest.objects.filter(id=request.POST.get('id'),
                                             status=LeaveRequest.STATUS_PENDING).first()
            if lr:
                ok = request.POST.get('decision') == 'approve'
                lr.status = LeaveRequest.STATUS_APPROVED if ok else LeaveRequest.STATUS_REJECTED
                lr.decided_by = request.user.username
                lr.decided_time = _now()
                lr.save()
                _notify_leave_decision(lr)
                messages.info(request, ('Đã DUYỆT' if ok else 'Đã TỪ CHỐI')
                              + f' đơn nghỉ của {lr.user}.')
            return redirect('/cham-cong-quan-ly')

    now = _now()
    month = request.GET.get('month') or now.strftime('%Y-%m')
    cfg = _att_cfg()
    today_str = now.strftime('%Y-%m-%d')
    summ, detail = {}, []
    device_users, noout = {}, []
    for r in Attendance.objects.filter(month=month).order_by('-day', 'user'):
        c = _att_calc(r, cfg)
        s = summ.setdefault(r.user, {'user': r.user, 'days': 0, 'hours': 0.0,
                                     'ot': 0.0, 'late': 0, 'fine': 0, 'ot_pay': 0})
        if r.check_in:
            s['days'] += 1
        s['hours'] += c['hours']; s['ot'] += c['ot_hours']
        s['late'] += c['late_min']; s['fine'] += c['fine']; s['ot_pay'] += c['ot_pay']
        detail.append({'day': _fmt_day(r.day), 'day_iso': r.day, 'wd': c['weekday'], 'user': r.user,
                       'in': _hm(r.check_in), 'out': _hm(r.check_out), 'hours': c['hours'],
                       'late': c['late_min'], 'ot': c['ot_hours'], 'fine': c['fine'],
                       'rest': not c['workday'], 'holiday': c['holiday']})
        if r.device_in:
            device_users.setdefault(r.device_in, set()).add(r.user)
        if r.check_in and not r.check_out and r.day != today_str:
            noout.append({'day': _fmt_day(r.day), 'user': r.user})
    # Nghỉ phép đã duyệt trong tháng -> cột "Nghỉ phép" (kể cả người chưa có dòng chấm công)
    from pha.extra_views import _approved_leave_days
    leave_map = _approved_leave_days(month)
    for u, ds in leave_map.items():
        summ.setdefault(u, {'user': u, 'days': 0, 'hours': 0.0,
                            'ot': 0.0, 'late': 0, 'fine': 0, 'ot_pay': 0})
    for s in summ.values():
        s['leave'] = len(leave_map.get(s['user'], ()))
    summary = sorted(summ.values(), key=lambda x: -x['hours'])
    for s in summary:
        s['hours'] = round(s['hours'], 1); s['ot'] = round(s['ot'], 1)
    totals = {k: round(sum(s[k] for s in summary), 1) if k in ('hours', 'ot')
              else sum(s[k] for s in summary)
              for k in ('days', 'hours', 'ot', 'late', 'fine', 'ot_pay', 'leave')}
    shared = [{'token': tok[:8] + '…', 'users': sorted(us)}
              for tok, us in device_users.items() if len(us) > 1]
    bindings = []
    for b in DeviceBind.objects.all().order_by('username'):
        try:
            when = (b.bound_time.astimezone(_VN) if _VN else b.bound_time).strftime('%d/%m %H:%M')
        except Exception:
            when = ''
        bindings.append({'user': b.username, 'when': when, 'ua': (b.user_agent or '')[:70]})
    months = sorted(set(Attendance.objects.values_list('month', flat=True)), reverse=True) \
        or [now.strftime('%Y-%m')]
    cur_ip = _client_ip(request)
    ips_set = bool(_attendance_ips())
    weekday_opts = [{'i': i, 'name': _WEEKDAYS[i], 'on': (i in cfg['workdays'])} for i in range(7)]
    from django.contrib.auth.models import User as _User
    emp_users = sorted(set(_User.objects.values_list('username', flat=True))
                       | set(Attendance.objects.values_list('user', flat=True)))
    return render(request, 'cham_cong_quan_ly.html', {
        'ips': AppSetting.get('ATTENDANCE_IPS', ''), 'cur_ip': cur_ip,
        'cur_private': _ip_private(cur_ip),
        'cur_ok': (_ip_allowed(cur_ip) if ips_set else None),
        'cfg': cfg, 'weekday_opts': weekday_opts,
        'month': month, 'month_label': _fmt_month(month),
        'months': [{'value': m, 'label': _fmt_month(m)} for m in months],
        'summary': summary, 'totals': totals, 'detail': detail[:300],
        'shared': shared, 'noout': noout[:50], 'bindings': bindings,
        'api_key': _api_key(),
        'api_base': request.scheme + '://' + request.get_host(),
        'ketoan_url': AppSetting.get('KETOAN_SALARY_URL', ''),
        'ketoan_pull_key': AppSetting.get('KETOAN_PULL_KEY', ''),
        'emp_users': emp_users,
        'today_iso': now.strftime('%Y-%m-%d'),
        'nudge_key': _nudge_key_for_view(),
        'api_base2': request.scheme + '://' + request.get_host(),
        'holidays': _upcoming_holidays(now, cfg),
        'prod_url': AppSetting.get('KETOAN_PROD_URL', ''),
        'prod_key': AppSetting.get('KETOAN_PROD_KEY', ''),
        'leave_pending': _leave_rows(status='pending'),
        'leave_recent': _leave_rows(exclude_pending=True, limit=10),
    })


def _leave_rows(status=None, exclude_pending=False, limit=50):
    """Danh sách đơn nghỉ phép cho trang quản lý (đã format ngày)."""
    from pha.models import LeaveRequest
    qs = LeaveRequest.objects.all()
    if status:
        qs = qs.filter(status=status)
    if exclude_pending:
        qs = qs.exclude(status=LeaveRequest.STATUS_PENDING)
    out = []
    for lr in qs[:limit]:
        out.append({'id': lr.id, 'user': lr.user,
                    'rng': _fmt_day(lr.from_day) + ('' if lr.from_day == lr.to_day
                                                    else ' → ' + _fmt_day(lr.to_day)),
                    'reason': lr.reason, 'status': lr.status,
                    'decided_by': lr.decided_by})
    return out


@csrf_exempt
@staff_required
def export_cham_cong_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from pha.models import Attendance
    now = _now()
    month = request.GET.get('month') or now.strftime('%Y-%m')
    wb = Workbook()
    head_fill = PatternFill('solid', fgColor='198754')
    head_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')
    cfg = _att_cfg()
    summ = {}
    ws2 = wb.active
    ws2.title = "Chi tiet"
    ws2.append(["Ngày", "Thứ", "Nhân viên", "Giờ vào", "Giờ ra", "Số giờ", "Đi muộn (phút)", "Tăng ca (giờ)", "Phạt (đ)"])
    for c in ws2[1]:
        c.fill = head_fill; c.font = head_font; c.alignment = center
    for r in Attendance.objects.filter(month=month).order_by('day', 'user'):
        c = _att_calc(r, cfg)
        s = summ.setdefault(r.user, {'days': 0, 'hours': 0.0, 'ot': 0.0, 'late': 0, 'fine': 0, 'ot_pay': 0})
        if r.check_in:
            s['days'] += 1
        s['hours'] += c['hours']; s['ot'] += c['ot_hours']
        s['late'] += c['late_min']; s['fine'] += c['fine']; s['ot_pay'] += c['ot_pay']
        ws2.append([_fmt_day(r.day), c['weekday'] + (' (nghỉ)' if not c['workday'] else ''),
                    r.user, _hm(r.check_in), _hm(r.check_out), c['hours'],
                    c['late_min'], c['ot_hours'], c['fine']])
    for col, w in zip('ABCDEFGHI', (10, 10, 16, 9, 9, 8, 13, 12, 12)):
        ws2.column_dimensions[col].width = w
    ws = wb.create_sheet("Tong hop", 0)
    ws.append([f"BẢNG CÔNG THÁNG {_fmt_month(month)}"])
    ws.append(["Nhân viên", "Ngày công", "Tổng giờ", "Tăng ca (giờ)", "Đi muộn (phút)", "Phạt (đ)", "Tiền tăng ca (đ)"])
    for c in ws[2]:
        c.fill = head_fill; c.font = head_font; c.alignment = center
    for u, s in sorted(summ.items(), key=lambda kv: -kv[1]['hours']):
        ws.append([u, s['days'], round(s['hours'], 1), round(s['ot'], 1), s['late'], s['fine'], s['ot_pay']])
    for col, w in zip('ABCDEFG', (18, 11, 11, 13, 14, 12, 16)):
        ws.column_dimensions[col].width = w
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="bang_cong_{month}.xlsx"'
    wb.save(resp)
    return resp


# ===================== XỬ LÝ ẢNH (tab cho chủ) =====================
def _fmt_name(filename):
    """Tên hiển thị gọn: '18:58 08/06 · Asset 2@4x.png' từ tên file có timestamp."""
    try:
        date_part, time_part, rest = filename.split('_', 2)
        return f'{time_part[:5].replace("-", ":")} {date_part[8:10]}/{date_part[5:7]} · {rest}'
    except (ValueError, IndexError):
        return filename


# Bộ nhớ tạm: chỉ giữ N kết quả gần nhất, xoá kết quả cũ hơn (kèm file ảnh) cho gọn.
RESULT_CACHE_KEEP = 10


def _prune_image_results(keep=RESULT_CACHE_KEEP):
    # KHÔNG xoá kết quả mới tạo trong 24h (khách /thiet-ke còn đang xem/tải ảnh
    # qua URL tuyệt đối; xoá sớm sẽ vỡ ảnh dù vừa trả về xong).
    from django.utils import timezone as _tz
    cutoff = _tz.now() - timedelta(hours=24)
    old_ids = list(ImageResult.objects.order_by('-created_time')
                   .values_list('id', flat=True)[keep:])
    if old_ids:
        old_ids = list(ImageResult.objects.filter(id__in=old_ids, created_time__lt=cutoff)
                       .values_list('id', flat=True))
    if not old_ids:
        return
    for obj in ImageResult.objects.filter(id__in=old_ids):
        for fn in (obj.name, obj.enhanced_name, obj.design_name, obj.name_output):
            if fn:
                try:
                    os.remove(os.path.join(settings.MEDIA_ROOT, fn))
                except OSError:
                    pass
    ImageResult.objects.filter(id__in=old_ids).delete()


def _get_img(request):
    data = request.GET.get('file_url') or request.POST.get('file_url')
    if not data:
        return None
    return ImageResult.objects.filter(name=data.replace('/media/', '')) \
        .order_by('-created_time').first()


def _colors_with_edits(res, request):
    """Trả về bản sao bảng màu, áp mã DALI người dùng vừa sửa (gửi qua 'edits',
    dạng {stt: 'MÃ'}). KHÔNG ghi vào bảng tham chiếu DALI và KHÔNG lưu DB —
    chỉ dùng cho file tải về (vì mã sửa chưa qua kiểm nghiệm)."""
    colors = [list(r) for r in (res.colors or [])]
    raw = request.GET.get('edits') or request.POST.get('edits')
    if raw:
        try:
            edits = json.loads(raw)
        except (ValueError, TypeError):
            edits = {}
        if isinstance(edits, dict):
            for row in colors:
                key = str(row[0])
                val = edits.get(key)
                if val:
                    while len(row) < 3:
                        row.append('')
                    row[2] = str(val).strip()
    return colors


# ===== PRESET TỰ LƯU (người dùng tự tạo gói cấu hình cho từng loại ảnh) =====
def _load_custom_presets():
    from pha.models import AppSetting
    try:
        return json.loads(AppSetting.get('IMAGE_PRESETS', '{}')) or {}
    except (ValueError, TypeError):
        return {}


def _save_custom_presets(d):
    from pha.models import AppSetting
    AppSetting.set('IMAGE_PRESETS', json.dumps(d, ensure_ascii=False))


def _all_presets_ui():
    """Gộp preset dựng sẵn + preset người dùng tự lưu (cho giao diện)."""
    from pha.ai_enhance import presets_for_ui
    out = {}
    for k, v in presets_for_ui().items():
        v = dict(v)
        v['custom'] = False
        v['base'] = k
        out[k] = v
    for k, v in _load_custom_presets().items():
        out[k] = {
            'label': v.get('label', k), 'desc': v.get('desc', 'Preset của tôi'),
            'color_limit': v.get('color_limit', 0), 'min_area': v.get('min_area', 0),
            'smooth': v.get('smooth', 0), 'enhance': bool(v.get('enhance')),
            'base': v.get('base', 'photo'), 'custom': True,
        }
    return out


def _resolve_preset_ai(preset_key):
    """Trả (ai_prompt, use_refs) cho preset (kể cả preset tự lưu)."""
    from pha.ai_enhance import get_preset
    custom = _load_custom_presets()
    if preset_key in custom:
        base = get_preset(custom[preset_key].get('base', 'photo'))
        return base.get('prompt'), bool(base.get('use_refs'))
    p = get_preset(preset_key)
    return p.get('prompt'), bool(p.get('use_refs'))


@csrf_exempt
@staff_required
def anh_preset(request):
    """Lưu / xoá preset tự tạo (gói cấu hình cho từng loại ảnh)."""
    if request.method != 'POST':
        return HttpResponseNotFound('POST only')
    action = request.POST.get('action')
    d = _load_custom_presets()
    if action == 'save':
        name = (request.POST.get('name') or '').strip()
        if not name:
            return JsonResponse({'ok': False, 'msg': 'Thiếu tên preset.'})

        def _i(k, lo, hi):
            try:
                return max(lo, min(hi, int(request.POST.get(k) or 0)))
            except ValueError:
                return 0
        d[name] = {
            'label': name + ' (của tôi)', 'desc': 'Preset tự lưu',
            'color_limit': _i('color_limit', 0, 250), 'min_area': _i('min_area', 0, 100000),
            'smooth': _i('smooth', 0, 3),
            'enhance': request.POST.get('enhance') in ('1', 'on', 'true'),
            'base': (request.POST.get('base') or 'photo').strip(),
        }
        _save_custom_presets(d)
        return JsonResponse({'ok': True, 'presets': _all_presets_ui(), 'selected': name})
    if action == 'delete':
        d.pop((request.POST.get('name') or '').strip(), None)
        _save_custom_presets(d)
        return JsonResponse({'ok': True, 'presets': _all_presets_ui()})
    return JsonResponse({'ok': False})


@csrf_exempt
@staff_required
def xu_ly_anh(request):
    from pha.imageproc import process_image
    from pha.ai_enhance import is_configured as ai_configured
    from pha import style_library

    from pha.ai_enhance import DEFAULT_PRESET

    def build_ctx():
        last = ImageResult.objects.all().order_by('-created_time')[:RESULT_CACHE_KEEP]
        return {'last_query': [{'name': _fmt_name(q.name), 'url': q.name} for q in last],
                'ai_available': ai_configured(),
                'style_categories': style_library.categories(),
                'presets_json': json.dumps(_all_presets_ui(), ensure_ascii=False),
                'default_preset': DEFAULT_PRESET}

    if request.method == 'POST' and request.FILES.get('image'):
        upload = request.FILES['image']
        fss = FileSystemStorage()
        # uuid + lấy tên THẬT fss.save trả về -> tên DUY NHẤT, record trỏ đúng file
        # (2 nhân viên/2 tab upload cùng giây + cùng tên file -> trước đây lẫn ảnh).
        name = f'{datetime.now():%Y-%m-%d_%H-%M-%S}_{uuid.uuid4().hex[:8]}_{upload.name}'
        name = fss.save(name, upload)
        enhance = request.POST.get('enhance') in ('1', 'on', 'true')
        style_category = (request.POST.get('style_category') or '').strip() or None
        try:
            color_limit = int(request.POST.get('color_limit') or 0)
        except ValueError:
            color_limit = 0
        color_limit = max(0, min(color_limit, 250))  # 0 = không giới hạn (trần 250 = LIMIT_NUM_COLOR)
        try:
            min_area = int(request.POST.get('min_area') or 0)
        except ValueError:
            min_area = 0
        min_area = max(0, min(min_area, 100000))  # 0 = không lọc mảng nhỏ
        try:
            smooth = int(request.POST.get('smooth') or 0)
        except ValueError:
            smooth = 0
        smooth = max(0, min(smooth, 3))  # 0=không, 1=nhẹ, 2=vừa, 3=mạnh
        preset_key = (request.POST.get('preset') or 'anime').strip()
        ai_prompt, use_refs = _resolve_preset_ai(preset_key)
        # Thanh "Mức độ AI" (1 lựa chọn = 1 lần gọi, KHÔNG nhân tiền): nếu chọn mức
        # thì DÙNG prompt mức đó thay prompt preset (nhe=giữ thật, vua, manh=nghệ
        # thuật, nen=nền tối giản). Để trống = theo loại tranh như cũ.
        ai_level = (request.POST.get('ai_level') or '').strip()
        if ai_level:
            try:
                from pha.ai_levels import LEVELS as _AILV
                _lvmap = {k: p for k, _lb, p in _AILV}
                if ai_level in _lvmap:
                    ai_prompt = _lvmap[ai_level]
                    use_refs = False           # mức tự chứa hướng dẫn, không cần ảnh mẫu
            except Exception:
                pass
        size_str = (request.POST.get('print_size') or '40x50').strip()
        try:
            dims = [int(x) for x in size_str.lower().replace(' ', '').split('x') if x]
            print_long_cm = max(dims) if dims else 0
        except ValueError:
            print_long_cm = 0
        # Chế độ CHI TIẾT (preset Cây/Hoa): giữ nhiều ô nhỏ, số nhỏ — ít gộp mảng.
        # face_priority: CHỈ preset chân dung 'photo' -> dò & bảo vệ ngũ quan.
        from pha.ai_enhance import get_preset as _gp
        _ps = _gp(preset_key)
        detail = bool(_ps.get('detail'))
        face_priority = bool(_ps.get('face_priority'))
        large = bool(_ps.get('large'))                 # TRANH KHỔ TO SIÊU CHI TIẾT
        # "Độ chi tiết đánh số" (chỉ preset chi tiết): hạ sàn cỡ số -> giữ ô nhỏ hơn.
        from pha.color_index_lib import detail_scale as _dscale
        num_detail_key = (request.POST.get('num_detail') or 'chuan').strip()
        num_detail = _dscale(num_detail_key)
        rec = ImageResult.objects.create(
            name=name, status=ImageResult.STATUS_PROCESSING, user=request.user.username,
            params={'enhance': enhance, 'color_limit': color_limit, 'min_area': min_area,
                    'smooth': smooth, 'style_category': style_category or '',
                    'preset': preset_key, 'print_size': size_str, 'ai_level': ai_level,
                    'detail': detail, 'face_priority': face_priority, 'large': large,
                    'num_detail': num_detail_key})
        _img_executor.submit(process_image, rec.id, name, enhance, style_category,
                             color_limit, min_area, smooth, ai_prompt, use_refs,
                             print_long_cm, detail, face_priority, large, num_detail)
        _prune_image_results()                 # giữ 10 kết quả gần nhất (bộ nhớ tạm)
        ctx = build_ctx()
        ctx['file_url'] = '/media/' + name
        return render(request, 'xu_ly_anh.html', ctx)
    _prune_image_results()
    return render(request, 'xu_ly_anh.html', build_ctx())


@csrf_exempt
@staff_required
def kho_mau(request):
    """Kho mẫu thành phẩm: tải hàng loạt + phân loại + xoá. Dùng làm ảnh tham
    chiếu phong cách cho AI khi tăng cường ảnh khách."""
    from pha import style_library
    from pha.models import StyleSample
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'upload':
            files = request.FILES.getlist('images')
            category = (request.POST.get('category') or '').strip()
            n = 0
            for f in files:
                if not f.content_type or not f.content_type.startswith('image/'):
                    continue
                try:
                    style_library.add_sample(f, category=category, user=request.user.username)
                    n += 1
                except Exception:
                    continue
            messages.info(request, f'Đã thêm {n} mẫu' + (f' vào nhãn "{category}".' if category else '.'))
        elif action == 'delete':
            sid = request.POST.get('id')
            obj = StyleSample.objects.filter(id=sid).first()
            if obj:
                try:
                    os.remove(os.path.join(settings.MEDIA_ROOT, obj.name))
                except OSError:
                    pass
                obj.delete()
                messages.info(request, 'Đã xoá 1 mẫu.')
        return redirect('/kho-mau')

    cat = (request.GET.get('cat') or '').strip()
    qs = StyleSample.objects.all()
    if cat:
        qs = qs.filter(category=cat)
    total = StyleSample.objects.count()
    items = list(qs[:300])
    return render(request, 'kho_mau.html', {
        'items': items, 'total': total, 'shown': len(items),
        'categories': style_library.category_options(), 'cur_cat': cat,
        'truncated': qs.count() > 300,
    })


def _mask_key(k):
    k = (k or '').strip()
    if not k:
        return ''
    if len(k) <= 8:
        return '••••'
    return k[:4] + '••••••' + k[-4:]


def _test_google_key():
    """Gọi nhẹ Google API để xác thực khoá (models.list — không tốn phí tạo ảnh)."""
    from pha.ai_enhance import get_api_key
    key = get_api_key()
    if not key:
        return False, 'Chưa có API key.'
    try:
        from google import genai
    except ImportError:
        return False, 'Máy chủ chưa cài thư viện google-genai (pip install google-genai).'
    try:
        client = genai.Client(api_key=key)
        next(iter(client.models.list()), None)
        return True, 'Kết nối Google AI thành công.'
    except Exception as e:
        return False, f'Khoá không dùng được: {e}'


def _test_openai_key():
    """Gọi nhẹ OpenAI để xác thực khoá (models.list — không tốn phí tạo ảnh)."""
    from pha.ai_enhance import get_openai_key
    key = get_openai_key()
    if not key:
        return False, 'Chưa có OpenAI API key.'
    try:
        from openai import OpenAI
    except ImportError:
        return False, 'Máy chủ chưa cài thư viện openai (pip install openai) — chạy update.sh.'
    try:
        client = OpenAI(api_key=key)
        next(iter(client.models.list()), None)
        return True, 'Kết nối OpenAI thành công.'
    except Exception as e:
        return False, f'Khoá không dùng được: {e}'


@csrf_exempt
@staff_required
def cai_dat_ai(request):
    """Nhập / lưu / kiểm tra Google API key cho tính năng tăng cường ảnh."""
    from pha.models import AppSetting
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'save':
            key = (request.POST.get('api_key') or '').strip()
            if not key:
                return JsonResponse({'ok': False, 'msg': 'Chưa nhập API key.'})
            AppSetting.set('GOOGLE_API_KEY', key)
            return JsonResponse({'ok': True, 'msg': 'Đã lưu thành công.', 'masked': _mask_key(key)})
        if action == 'clear':
            AppSetting.objects.filter(key='GOOGLE_API_KEY').delete()
            return JsonResponse({'ok': True, 'msg': 'Đã xoá khoá đã lưu.'})
        if action == 'test':
            ok, msg = _test_google_key()
            return JsonResponse({'ok': ok, 'msg': msg})
        # ---- OpenAI (ChatGPT) + chọn nhà cung cấp AI ----
        if action == 'save_openai':
            k = (request.POST.get('api_key') or '').strip()
            if not k:
                return JsonResponse({'ok': False, 'msg': 'Chưa nhập OpenAI key.'})
            AppSetting.set('OPENAI_API_KEY', k)
            return JsonResponse({'ok': True, 'msg': 'Đã lưu khoá OpenAI.', 'masked': _mask_key(k)})
        if action == 'clear_openai':
            AppSetting.objects.filter(key='OPENAI_API_KEY').delete()
            return JsonResponse({'ok': True, 'msg': 'Đã xoá khoá OpenAI.'})
        if action == 'test_openai':
            ok, msg = _test_openai_key()
            return JsonResponse({'ok': ok, 'msg': msg})
        if action == 'save_provider':
            prov = (request.POST.get('provider') or 'gemini').strip().lower()
            if prov not in ('gemini', 'openai'):
                prov = 'gemini'
            AppSetting.set('AI_PROVIDER', prov)
            AppSetting.set('AI_FALLBACK', '1' if request.POST.get('fallback') == '1' else '0')
            return JsonResponse({'ok': True, 'msg': 'Đã lưu: nhà chính = %s%s.'
                                 % (prov.upper(), ' + dự phòng' if request.POST.get('fallback') == '1' else '')})
        # ---- Khoá API cho web bán hàng (tranhdali.vn/thiet-ke) ----
        if action == 'save_thietke':
            k = (request.POST.get('api_key') or '').strip()
            if not k:
                return JsonResponse({'ok': False, 'msg': 'Chưa nhập khoá.'})
            AppSetting.set('THIETKE_API_KEY', k)
            return JsonResponse({'ok': True, 'msg': 'Đã lưu khoá Thiết kế.', 'masked': _mask_key(k), 'key': k})
        if action == 'gen_thietke':
            import secrets
            k = 'tk_' + secrets.token_urlsafe(24)
            AppSetting.set('THIETKE_API_KEY', k)
            return JsonResponse({'ok': True, 'msg': 'Đã tạo & lưu khoá mới.', 'masked': _mask_key(k), 'key': k})
        if action == 'clear_thietke':
            AppSetting.objects.filter(key='THIETKE_API_KEY').delete()
            return JsonResponse({'ok': True, 'msg': 'Đã xoá khoá Thiết kế.'})
        return JsonResponse({'ok': False, 'msg': 'Hành động không hợp lệ.'})

    db_key = (AppSetting.get('GOOGLE_API_KEY') or '').strip()
    env_key = os.environ.get('GOOGLE_API_KEY') or os.environ.get('GEMINI_API_KEY') or ''
    cur = db_key or env_key
    tk_db = (AppSetting.get('THIETKE_API_KEY') or '').strip()
    tk_env = os.environ.get('THIETKE_API_KEY') or ''
    tk_cur = tk_db or tk_env
    oa_db = (AppSetting.get('OPENAI_API_KEY') or '').strip()
    oa_env = os.environ.get('OPENAI_API_KEY') or ''
    oa_cur = oa_db or oa_env
    from pha import ai_enhance as _ai
    return render(request, 'cai_dat_ai.html', {
        'has_key': bool(cur),
        'masked': _mask_key(cur),
        'from_env': bool(not db_key and env_key),
        'tk_has': bool(tk_cur),
        'tk_masked': _mask_key(tk_cur),
        'tk_from_env': bool(not tk_db and tk_env),
        'oa_has': bool(oa_cur),
        'oa_masked': _mask_key(oa_cur),
        'oa_from_env': bool(not oa_db and oa_env),
        'provider': _ai.get_provider(),
        'fallback': _ai.get_fallback(),
    })


@csrf_exempt
@staff_required
def anh_result(request):
    from pha.imageproc import split_list, mark_if_stuck
    res = _get_img(request)
    if not res:
        return JsonResponse({'status': 'processing'})
    if res.status == ImageResult.STATUS_PROCESSING:
        if not mark_if_stuck(res):          # kẹt quá lâu -> rơi xuống nhánh lỗi
            return JsonResponse({'status': 'processing',
                                 'progress': (res.params or {}).get('progress')})
    if res.status == ImageResult.STATUS_ERROR:
        return JsonResponse({'status': 'error', 'error': res.error_message})
    return JsonResponse({'status': 'done', 'img_output': '/media/' + res.name_output,
                         'original': '/media/' + res.name,
                         'enhanced': ('/media/' + res.enhanced_name) if res.enhanced_name else '',
                         'design': ('/media/' + res.design_name) if res.design_name else '',
                         'params': res.params or {}, 'warn': res.error_message or '',
                         'colors': split_list(10, res.colors)})


@csrf_exempt
@staff_required
def anh_save_color(request):
    """Sửa mã DALI cho 1 màu: ghi vào bảng tham chiếu DALI + cập nhật ảnh hiện tại."""
    from pha import dali_match
    if request.method != 'POST':
        return HttpResponseNotFound('POST only')
    hex_value = request.POST.get('hex', '')
    dali = request.POST.get('dali', '')
    stt = request.POST.get('stt', '')
    file_url = request.POST.get('file_url', '')
    ok, msg = dali_match.add_entry(hex_value, dali)
    if not ok:
        return JsonResponse({'ok': False, 'msg': msg})
    if file_url:
        res = ImageResult.objects.filter(name=file_url.replace('/media/', '')) \
            .order_by('-created_time').first()
        if res and res.colors:
            changed = False
            for row in res.colors:
                if str(row[0]) == str(stt) and len(row) > 2:
                    row[2] = dali; changed = True
            if changed:
                res.save()
    return JsonResponse({'ok': True, 'msg': msg})


@csrf_exempt
@staff_required
def anh_nearest_dali(request):
    """Trả mã DALI gần nhất cho 1 mã HEX (dùng khi đổi màu trực tiếp ở bảng màu)."""
    from pha import dali_match
    h = (request.GET.get('hex') or '').strip().lstrip('#')
    try:
        rgb = (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
    except (ValueError, IndexError):
        return JsonResponse({'dali': ''})
    return JsonResponse({'dali': dali_match.nearest_dali(rgb)})


@csrf_exempt
@staff_required
def anh_export_colors(request):
    res = _get_img(request)
    if not res:
        return HttpResponseNotFound('no result')
    resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    resp['Content-Disposition'] = 'attachment; filename="bang_mau_dali.csv"'
    resp.write('﻿')
    w = csv.writer(resp)
    w.writerow(['STT', 'HEX', 'R', 'G', 'B', 'Ma_DALI', 'Phan_tram'])
    for row in _colors_with_edits(res, request):
        h = row[1].lstrip('#')
        try:
            r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        except (ValueError, IndexError):
            r = g = b = ''
        w.writerow([row[0], row[1], r, g, b, row[2] if len(row) > 2 else '', row[3] if len(row) > 3 else ''])
    return resp


@csrf_exempt
@staff_required
def anh_export_xlsx(request):
    from pha.exports import build_xlsx
    res = _get_img(request)
    if not res:
        return HttpResponseNotFound('no result')
    out = os.path.join(settings.MEDIA_ROOT, 'bang_mau_dali.xlsx')
    build_xlsx(_colors_with_edits(res, request), out)
    with open(out, 'rb') as f:
        data = f.read()
    resp = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="bang_mau_dali.xlsx"'
    return resp


@csrf_exempt
@staff_required
def anh_legend(request):
    from pha.exports import build_legend_image
    res = _get_img(request)
    if not res:
        return HttpResponseNotFound('no result')
    left = res.name if os.path.exists(os.path.join(settings.MEDIA_ROOT, res.name)) else res.name_output
    left_path = os.path.join(settings.MEDIA_ROOT, left)
    title = (request.GET.get('title') or '').strip()
    out = os.path.join(settings.MEDIA_ROOT, f'{res.name_output or res.name}_legend.png')
    build_legend_image(left_path, _colors_with_edits(res, request), out, title=title)
    with open(out, 'rb') as f:
        data = f.read()
    resp = HttpResponse(data, content_type='image/png')
    resp['Content-Disposition'] = 'attachment; filename="bang_mau_DALI.png"'
    return resp


@csrf_exempt
@staff_required
def anh_download_result(request):
    from pha.imageproc import get_paint_image
    result_url = request.GET.get('result_url')
    image_name = request.GET.get('image_name') or 'result'
    scale = request.GET.get('scale_option') or '20x20'
    orientation = request.GET.get('orientation_option') or 'auto'
    try:
        file_paint, file_a3 = get_paint_image(result_url, image_name, scale, orientation)
        return JsonResponse({'file_paint': file_paint, 'file_a3': file_a3, 'origin_result': result_url})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

