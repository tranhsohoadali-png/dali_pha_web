"""
Các tiện ích xuất file cho chương trình gộp:
- build_xlsx: xuất bảng màu ra Excel (.xlsx), mỗi dòng có ô tô màu nền theo HEX.
- build_legend_image: dựng ảnh bản đồ màu kèm bảng chú giải số thứ tự -> mã DALI.

Phần Excel dùng openpyxl (đã copy vào python_embed). Phần ảnh dùng PIL/numpy
(có sẵn). Không phụ thuộc pandas.
"""
import csv
import io
import os
import re

import numpy as np
from PIL import Image, ImageDraw, ImageFont

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def parse_color_file(django_file, filename=''):
    """
    Đọc file Excel (.xlsx) hoặc CSV chứa bảng màu DALI, trả về list (hex, dali).

    - Tự nhận cột theo tiêu đề: cột có chữ 'hex' -> mã HEX, có 'dali' -> mã DALI.
    - Nếu không có tiêu đề rõ ràng: lấy cột 1 = HEX, cột 2 = DALI.
    - HEX chấp nhận '1a2b3c', '#1a2b3c', hoặc số bị mất số 0 đầu (tự bù đủ 6).
    """
    name = (filename or getattr(django_file, 'name', '')).lower()

    rows = []
    if name.endswith('.csv'):
        data = django_file.read()
        if isinstance(data, bytes):
            data = data.decode('utf-8-sig', errors='replace')
        for r in csv.reader(io.StringIO(data)):
            rows.append(r)
    else:  # .xlsx / .xlsm
        wb = load_workbook(django_file, read_only=True, data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r))

    if not rows:
        return []

    def clean_hex(v):
        if v is None:
            return None
        s = str(v).strip()
        if s.endswith('.0'):
            s = s[:-2]
        s = s.lstrip('#')
        if not re.fullmatch(r'[0-9a-fA-F]{1,6}', s):
            return None
        return s.zfill(6).lower()

    # ---- Nhận diện cột HEX / DALI theo tiêu đề ----
    col_hex, col_dali = None, None
    header_has_label = False
    for i, cell in enumerate(rows[0]):
        s = str(cell).strip().lower() if cell is not None else ''
        if s:
            header_has_label = True
        if col_hex is None and 'hex' in s:
            col_hex = i
        if col_dali is None and 'dali' in s:
            col_dali = i

    has_header = header_has_label
    data_rows = rows[1:] if has_header else rows
    ncols = max((len(r) for r in rows), default=0)

    # ---- Đoán cột còn thiếu bằng nội dung ----
    # Cột HEX: nhiều giá trị hợp lệ HEX nhất. Cột DALI: cột (không phải HEX)
    # có nhiều giá trị KHÁC NHAU nhất (mã DALI), bỏ qua cột rỗng / 1 giá trị.
    if col_hex is None or col_dali is None:
        sample = data_rows[:300]
        hex_score = [0] * ncols
        distinct = [set() for _ in range(ncols)]
        for r in sample:
            for c in range(min(len(r), ncols)):
                v = r[c]
                if clean_hex(v):
                    hex_score[c] += 1
                if v is not None and str(v).strip():
                    distinct[c].add(str(v).strip())
        if col_hex is None and ncols:
            col_hex = max(range(ncols), key=lambda c: hex_score[c])
        if col_dali is None and ncols:
            cand = [c for c in range(ncols) if c != col_hex and len(distinct[c]) > 1]
            if cand:
                col_dali = max(cand, key=lambda c: len(distinct[c]))
            else:
                col_dali = 0 if col_hex != 0 else 1

    if col_hex is None:
        col_hex = 0
    if col_dali is None:
        col_dali = 1 if col_hex == 0 else 0
    pairs = []
    for r in data_rows:
        if col_hex >= len(r) or col_dali >= len(r):
            continue
        h = clean_hex(r[col_hex])
        d = r[col_dali]
        d = '' if d is None else str(d).strip()
        if d.endswith('.0'):
            d = d[:-2]
        if not h or not d or d.lower() == 'nan':
            continue
        pairs.append((h, d))
    return pairs


def _hex_clean(hex_value):
    return str(hex_value).lstrip('#').upper()


def _hex_to_rgb(hex_value):
    h = _hex_clean(hex_value)
    try:
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except (ValueError, IndexError):
        return 0, 0, 0


def build_xlsx(colors, out_path):
    """
    colors: list các dòng [stt, "#HEX", dali, percent]
    Tạo file Excel với cột: STT | Màu | HEX | R | G | B | Mã DALI | % diện tích
    Ô cột "Màu" được tô nền đúng màu HEX.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bang mau DALI"

    headers = ['STT', 'Màu', 'HEX', 'R', 'G', 'B', 'Mã DALI', '% diện tích']
    widths = [6, 10, 12, 6, 6, 6, 14, 12]
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (title, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2E7D32')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
        ws.column_dimensions[c.column_letter].width = w

    for i, row in enumerate(colors, start=2):
        stt = row[0]
        hex_value = row[1] if len(row) > 1 else ''
        dali = row[2] if len(row) > 2 else ''
        percent = row[3] if len(row) > 3 else ''
        r, g, b = _hex_to_rgb(hex_value)
        clean = _hex_clean(hex_value)

        values = [stt, '', clean, r, g, b, dali, percent]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
        # tô màu ô "Màu"
        ws.cell(row=i, column=2).fill = PatternFill('solid', fgColor=clean.zfill(6))
        ws.row_dimensions[i].height = 22

    wb.save(out_path)
    return out_path


def _load_font(size):
    # Thử font Windows trước (máy tính), rồi font Linux (VPS) - đều scale theo size.
    for path in (
        r'C:\Windows\Fonts\arial.ttf',
        r'C:\Windows\Fonts\segoeui.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf',
    ):
        try:
            return ImageFont.truetype(path, size)
        except OSError:
            continue
    # Fallback cuối: font mặc định - thử truyền size (Pillow >=10 mới scale được).
    try:
        return ImageFont.load_default(size)
    except TypeError:
        return ImageFont.load_default()


def build_legend_image(left_image_path, colors, out_path, title='', max_per_col=12):
    """
    Dựng ảnh tải về theo bố cục mẫu ("K496" / "H092 TEST"):
      - Góc trên-trái: ảnh, phía dưới là tên mã (title) chữ lớn.
      - Bên phải: bảng chú giải, tự ĐỘNG CHIA NHIỀU CỘT khi nhiều màu
        (mỗi cột tối đa max_per_col màu). Mỗi dòng gồm
        [số thứ tự] · [ô màu] · [mã DALI].
    colors: [stt, "#HEX", dali, percent]
    """
    n = max(1, len(colors))

    # ---- Kích thước cố định để chú giải luôn rõ, không phụ thuộc số màu ----
    pad = 50
    gap = 48
    row_h = 132
    swatch_w = 300
    swatch_h = 88
    num_col_w = 150
    dali_w = 440
    col_w = num_col_w + swatch_w + 28 + dali_w

    num_font = _load_font(88)
    dali_font = _load_font(82)
    title_font = _load_font(150)

    ncols = (n + max_per_col - 1) // max_per_col
    height_rows = min(n, max_per_col)
    legend_h = height_rows * row_h

    # ---- Ảnh bên trái ----
    base = Image.open(left_image_path).convert('RGB')
    bw, bh = base.size
    img_box_w = 820
    scale = img_box_w / bw
    if bh * scale > legend_h:          # không cao quá khối chú giải
        scale = legend_h / bh
    bw, bh = int(bw * scale), int(bh * scale)
    base = base.resize((bw, bh))
    title_h = 210

    canvas_w = pad + bw + gap + ncols * col_w + (ncols - 1) * gap + pad
    canvas_h = pad + max(legend_h, bh + title_h) + pad
    canvas = Image.new('RGB', (canvas_w, canvas_h), 'white')
    canvas.paste(base, (pad, pad))

    draw = ImageDraw.Draw(canvas)
    if title:
        draw.text((pad + bw / 2, pad + bh + title_h * 0.5), str(title),
                  fill='black', font=title_font, anchor='mm')

    legend_x0 = pad + bw + gap
    for i, row in enumerate(colors):
        stt = row[0]
        hex_value = row[1] if len(row) > 1 else '#000000'
        dali = row[2] if len(row) > 2 else ''
        col = i // max_per_col
        r = i % max_per_col
        cx = legend_x0 + col * (col_w + gap)
        cy = pad + row_h * (r + 0.5)

        # số thứ tự
        draw.text((cx + num_col_w * 0.5, cy), str(stt), fill='black', font=num_font, anchor='mm')
        # ô màu
        sx0 = cx + num_col_w
        sy0 = cy - swatch_h / 2
        draw.rectangle([sx0, sy0, sx0 + swatch_w, sy0 + swatch_h],
                       fill=_hex_to_rgb(hex_value), outline=(180, 180, 180))
        # mã DALI
        draw.text((sx0 + swatch_w + 28, cy), str(dali), fill='black', font=dali_font, anchor='lm')

    canvas.save(out_path)
    return out_path
